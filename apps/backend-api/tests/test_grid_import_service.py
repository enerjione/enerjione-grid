"""grid_import_service — sablon uret + parse + apply (in-memory SQLite).

Kolon sirasi (COLUMNS) — Cihaz_Sira YOK:
  Bolge_Kodu, Bolge_Adi, Hat_Kodu, Hat_Adi, Direk_Sira, Direk_Adi,
  Enlem, Boylam, Direk_Tipi, Cihaz, Yon, Bransman_Hat_Kodu, Bransman_Direk_Sira

Yeni yapi: cihazlar direk satirlarinda DEGIL, ayri CIHAZ ara-satirlarinda
(Direk_Sira bos, Cihaz dolu). Ara-satir bir onceki direge baglanir.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.base import Base
from app.models.device import Device
from app.models.grid_topology import Line, LineSegment, Pole, Region  # noqa: F401
from app.services import grid_import_service as g


@pytest.fixture()
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    s = Session()
    yield s
    s.close()


def _row(bolge_k="", bolge_a="", hat_k="", hat_a="", sira=None, direk_a="",
         lat=None, lon=None, tip="", br_hat="", br_sira=""):
    """DIREK satiri (13 kolon). Cihaz/Yon bos."""
    return [bolge_k, bolge_a, hat_k, hat_a, sira, direk_a, lat, lon, tip,
            "", "", br_hat, br_sira]


def _dev_row(cihaz, yon=""):
    """CIHAZ ara-satiri: Direk_Sira bos, Cihaz dolu. Ustteki direge baglanir."""
    return ["", "", "", "", None, "", None, None, "", cihaz, yon, "", ""]


def _sheet(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Topoloji"
    ws.append(g.COLUMNS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_build_template_has_sheets_and_headers(db):
    buf = g.build_template_workbook(db)
    wb = load_workbook(buf)
    assert "Topoloji" in wb.sheetnames
    assert "Cihazlar" in wb.sheetnames
    assert "Hatlar" in wb.sheetnames   # bransman bagimli dropdown kaynagi
    assert "Yardim" in wb.sheetnames
    header = [c.value for c in wb["Topoloji"][1]]
    assert header == g.COLUMNS
    assert "Cihaz_Sira" not in header   # kaldirildi
    assert "Yon" in header and "Cihaz" in header
    assert len(header) == 14  # +Enerji_Rolu


def test_build_template_fills_existing_data_as_tree(db):
    """Sablon mevcut topolojiyi DOLU + agac (forward-fill: alt satirda bolge bos)."""
    reg = Region(code="R1", name="Bolge 1")
    db.add(reg); db.flush()
    line = Line(region_id=reg.id, code="L1", name="Hat 1"); db.add(line); db.flush()
    p1 = Pole(line_id=line.id, sequence_no=1, latitude=40.0, longitude=29.0, pole_type="pole")
    p2 = Pole(line_id=line.id, sequence_no=2, latitude=40.1, longitude=29.1, pole_type="pole")
    db.add_all([p1, p2]); db.flush()
    db.commit()

    wb = load_workbook(g.build_template_workbook(db))
    ws = wb["Topoloji"]
    # 1. veri satiri bolge dolu, 2. satir bolge bos (agac girinti).
    assert ws.cell(row=2, column=1).value == "R1"
    assert ws.cell(row=3, column=1).value in (None, "")   # forward-fill: bos


def test_build_template_has_conditional_formatting(db):
    """Tekrar eden cihaz icin kosullu bicim kurali Cihaz kolonuna eklenmis."""
    db.add(Device(code="D1", name="c1", ip_address="1.1.1.1", latitude=0, longitude=0))
    db.commit()
    wb = load_workbook(g.build_template_workbook(db))
    ws = wb["Topoloji"]
    # En az bir CF kurali olmali (Cihaz kolonu COUNTIF>1).
    ranges = list(ws.conditional_formatting)
    assert len(ranges) >= 1


def test_build_template_defines_line_named_ranges(db):
    """Her hat icin HAT_<token> named range tanimli (bagimli dropdown kaynagi)."""
    reg = Region(code="R1", name="B1"); db.add(reg); db.flush()
    ln = Line(region_id=reg.id, code="F-1", name="Fider 1"); db.add(ln); db.flush()
    db.add_all([
        Pole(line_id=ln.id, sequence_no=1, latitude=40.0, longitude=29.0, pole_type="pole"),
        Pole(line_id=ln.id, sequence_no=2, latitude=40.1, longitude=29.1, pole_type="pole"),
    ]); db.commit()
    wb = load_workbook(g.build_template_workbook(db))
    # 'F-1' -> HAT_F_1 (tire altcizgiye)
    assert "HAT_F_1" in wb.defined_names


def test_sanitize_range_token():
    assert g._range_name("F-1") == "HAT_F_1"
    assert g._range_name("E1") == "HAT_E1"
    assert g._range_name("HAT 2") == "HAT_HAT_2"
    # rakamla baslarsa L_ prefix
    assert g._range_name("1A").startswith("HAT_L_")


def test_parse_device_in_between_row(db):
    """Cihaz ara-satiri: ustteki direge baglanir, koordinat direk satirindan."""
    db.add(Device(code="DEV-1", name="Cihaz 1", ip_address="10.0.0.1", latitude=0, longitude=0))
    db.commit()
    rows = [
        _row("R1", "Bolge 1", "L1", "Hat 1", 1, "Bas", 40.0, 29.0, "transformer"),
        _dev_row("DEV-1"),                          # 1. direk sonrasi cihaz
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
    ]
    plan = g.parse_and_plan(_sheet(rows), db)
    c = plan.counts()
    assert c["regions"] == 1
    assert c["lines"] == 1
    assert c["poles"] == 2
    assert c["devices"] == 1
    assert c["errors"] == 0


def test_parse_invalid_and_unknown(db):
    db.add(Device(code="DEV-1", name="Cihaz 1", ip_address="10.0.0.1", latitude=0, longitude=0))
    db.commit()
    rows = [
        _row("R1", "Bolge 1", "L1", "Hat 1", 1, "Bas", 40.0, 29.0, "pole"),
        _dev_row("DEV-1"),
        _row(sira=2, lat=999.0, lon=29.2, tip="pole"),   # gecersiz enlem
        _dev_row("YOK"),                                 # bilinmeyen cihaz (2. direge)
    ]
    plan = g.parse_and_plan(_sheet(rows), db)
    c = plan.counts()
    assert c["poles"] == 1            # sadece direk 1 (direk 2 enlem hatali)
    assert c["devices"] == 1          # DEV-1
    assert c["errors"] == 2           # gecersiz enlem + bilinmeyen cihaz


def test_parse_multi_device_slot_auto_order(db):
    """Iki direk arasi 2 cihaz ara-satiri -> sira otomatik (satir sirasi)."""
    db.add(Device(code="D1", name="c1", ip_address="1.1.1.1", latitude=0, longitude=0))
    db.add(Device(code="D2", name="c2", ip_address="1.1.1.2", latitude=0, longitude=0))
    db.commit()
    rows = [
        _row("R1", "Bolge 1", "L1", "Hat 1", 1, "P1", 40.0, 29.0, "pole"),
        _dev_row("D1", "yesil"),   # slot 1->2, sira 1
        _dev_row("D2", "kirmizi"), # slot 1->2, sira 2 (otomatik)
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
    ]
    plan = g.parse_and_plan(_sheet(rows), db)
    assert plan.counts()["devices"] == 2
    assert plan.counts()["errors"] == 0
    result = g.apply_plan(plan, db); db.commit()
    assert result.segments_created == 2
    segs = db.query(LineSegment).order_by(LineSegment.device_position_t).all()
    assert [s.device_orientation for s in segs] == ["green_forward", "red_forward"]
    assert segs[0].device_position_t < segs[1].device_position_t   # sira korunur


def test_parse_invalid_direction(db):
    db.add(Device(code="D1", name="c1", ip_address="1.1.1.1", latitude=0, longitude=0)); db.commit()
    rows = [
        _row("R1", "B", "L1", "H", 1, "", 40.0, 29.0, "pole"),
        _dev_row("D1", "mavi"),   # gecersiz yon
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
    ]
    plan = g.parse_and_plan(_sheet(rows), db)
    assert any("Yon" in e.message for e in plan.errors)


def test_duplicate_device_error(db):
    db.add(Device(code="D1", name="c1", ip_address="1.1.1.1", latitude=0, longitude=0)); db.commit()
    rows = [
        _row("R1", "B", "L1", "H", 1, "", 40.0, 29.0, "pole"),
        _dev_row("D1"),
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
        _dev_row("D1"),   # ayni cihaz ikinci kez
    ]
    plan = g.parse_and_plan(_sheet(rows), db)
    assert any("birden fazla" in e.message for e in plan.errors)


def test_device_without_pole_errors(db):
    """Ustunde direk olmayan cihaz ara-satiri -> hata."""
    db.add(Device(code="D1", name="c1", ip_address="1.1.1.1", latitude=0, longitude=0)); db.commit()
    rows = [
        _row("R1", "B", "L1", "H", 1, "", 40.0, 29.0, "pole"),
        _dev_row("D1"),
    ]
    # Hat basi sonrasi ilk satir direk, cihaz ona baglanir -> hata YOK burada.
    # Ama hat basinda direkten ONCE cihaz gelirse hata olmali:
    rows2 = [
        ["R1", "B", "L1", "H", None, "", None, None, "", "D1", "", "", ""],  # direk yok, cihaz var
        _row(sira=1, lat=40.0, lon=29.0, tip="pole"),
    ]
    plan = g.parse_and_plan(_sheet(rows2), db)
    assert any("bağlı değil" in e.message or "direk" in e.message.lower() for e in plan.errors)


def test_apply_creates_topology(db):
    db.add(Device(code="DEV-1", name="Cihaz 1", ip_address="10.0.0.1", latitude=0, longitude=0)); db.commit()
    rows = [
        _row("R1", "Bolge 1", "L1", "Hat 1", 1, "P1", 40.0, 29.0, "pole"),
        _dev_row("DEV-1"),
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
    ]
    result = g.apply_plan(g.parse_and_plan(_sheet(rows), db), db); db.commit()
    assert result.regions_created == 1
    assert result.lines_created == 1
    assert result.poles_created == 2
    assert result.segments_created == 1
    assert db.query(Region).count() == 1


def test_apply_is_idempotent_upsert(db):
    rows = [
        _row("R1", "Bolge 1", "L1", "Hat 1", 1, "P1", 40.0, 29.0, "pole"),
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
    ]
    data = _sheet(rows)
    g.apply_plan(g.parse_and_plan(data, db), db); db.commit()
    r2 = g.apply_plan(g.parse_and_plan(data, db), db); db.commit()
    assert r2.regions_created == 0
    assert r2.poles_created == 0
    assert r2.poles_updated == 2
    assert db.query(Pole).count() == 2


def test_import_template_round_trip_keeps_existing_device(db):
    """Sablonu uret -> geri yukle: mevcut cihaz ara-satirda gelir, tekrar baglanmaz."""
    dev = Device(code="D1", name="c1", ip_address="1.1.1.1", latitude=0, longitude=0)
    db.add(dev); db.flush()
    reg = Region(code="R1", name="Bolge 1"); db.add(reg); db.flush()
    line = Line(region_id=reg.id, code="L1", name="Hat 1"); db.add(line); db.flush()
    p1 = Pole(line_id=line.id, sequence_no=1, latitude=40.0, longitude=29.0, pole_type="pole")
    p2 = Pole(line_id=line.id, sequence_no=2, latitude=40.1, longitude=29.1, pole_type="pole")
    db.add_all([p1, p2]); db.flush()
    db.add(LineSegment(line_id=line.id, from_pole_id=p1.id, to_pole_id=p2.id, device_id=dev.id))
    db.commit()

    data = g.build_template_workbook(db).getvalue()
    plan = g.parse_and_plan(data, db)
    result = g.apply_plan(plan, db); db.commit()

    assert plan.counts()["errors"] == 0
    assert plan.counts()["devices"] == 1   # cihaz ara-satirda geldi
    assert result.errors == []
    assert result.segments_created == 0    # zaten bagli, tekrar yaratmaz
    assert db.query(LineSegment).count() == 1


def test_branch_round_trip(db):
    """Bransmanli hat: sablon round-trip hatasiz, branched_from korunur."""
    reg = Region(code="R1", name="B1"); db.add(reg); db.flush()
    main = Line(region_id=reg.id, code="E1", name="Ana"); db.add(main); db.flush()
    mp1 = Pole(line_id=main.id, sequence_no=1, latitude=40.0, longitude=29.0, pole_type="pole")
    mp2 = Pole(line_id=main.id, sequence_no=2, latitude=40.1, longitude=29.1, pole_type="pole")
    db.add_all([mp1, mp2]); db.flush()
    branch = Line(region_id=reg.id, code="E2", name="Dal", branched_from_pole_id=mp2.id)
    db.add(branch); db.flush()
    db.add_all([
        Pole(line_id=branch.id, sequence_no=1, latitude=41.0, longitude=30.0, pole_type="pole"),
        Pole(line_id=branch.id, sequence_no=2, latitude=41.1, longitude=30.1, pole_type="pole"),
    ]); db.commit()

    data = g.build_template_workbook(db).getvalue()
    plan = g.parse_and_plan(data, db)
    assert plan.counts()["errors"] == 0
    # Bransman bilgisi plana tasindi
    branch_line = plan.lines.get("R1|E2")
    assert branch_line is not None
    assert branch_line.branch_line_code == "E1"
    assert branch_line.branch_pole_seq == 2

    result = g.apply_plan(plan, db); db.commit()
    assert result.errors == []
    refreshed = db.query(Line).filter(Line.code == "E2").one()
    assert refreshed.branched_from_pole_id == mp2.id


def test_same_line_code_in_different_regions_does_not_merge_poles(db):
    rows = [
        _row("R1", "Bolge 1", "L1", "Hat 1", 1, "R1P1", 40.0, 29.0, "pole"),
        _row(sira=2, lat=40.1, lon=29.1, tip="pole"),
        _row("R2", "Bolge 2", "L1", "Hat 1", 1, "R2P1", 41.0, 30.0, "pole"),
        _row(sira=2, lat=41.1, lon=30.1, tip="pole"),
    ]
    g.apply_plan(g.parse_and_plan(_sheet(rows), db), db); db.commit()

    assert db.query(Region).count() == 2
    assert db.query(Line).count() == 2
    assert db.query(Pole).count() == 4


# --------------------------------------------------------------------------- #
# Hizli Yapistir + sihirbaz
# --------------------------------------------------------------------------- #
def _quick_sheet(rows: list[list], with_topo: bool = False) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Topoloji"
    ws0.append(g.COLUMNS)
    if with_topo:
        ws0.append(_row("B1", "Bolge 1", "H1", "Hat 1", 1, "", 39.0, 32.0))
        ws0.append(_row(sira=2, lat=39.1, lon=32.1))
    ws = wb.create_sheet(g.QUICK_SHEET)
    ws.append(g.QUICK_COLUMNS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_coord_pair_bicimleri():
    """Virgul, bosluk, noktali virgul, ondalik virgul ve TERS yapistirma."""
    assert g._parse_coord_pair("39.92042, 32.85411") == (39.92042, 32.85411)
    assert g._parse_coord_pair("39.92 32.85") == (39.92, 32.85)
    assert g._parse_coord_pair("39.92;32.85") == (39.92, 32.85)
    assert g._parse_coord_pair("39,92 32,85") == (39.92, 32.85)
    # Ters (boylam once): |lat|>90 -> otomatik takas.
    assert g._parse_coord_pair("132.85, 39.92") == (39.92, 132.85)
    assert g._parse_coord_pair("") is None
    assert g._parse_coord_pair("abc") is None


def test_quick_sheet_otomatik_sira_ve_forward_fill(db):
    """Bolge/Hat bir kez yazilir; sira otomatik 1..N; 'ornek' satiri atlanir."""
    data = _quick_sheet([
        ["B1", "H1", "ornek: 39.9, 32.8", "", ""],
        ["B1", "H1", "39.0, 32.0", "", ""],
        ["", "", "39.1, 32.1", "ara direk", ""],
        ["", "", "39.2 32.2", "", "transformer"],
    ])
    plan = g.parse_and_plan(data, db)
    assert [e.message for e in plan.errors] == []
    assert [(p.sequence_no, p.topology_role, p.energy_role) for p in plan.poles] == [
        (1, "transit", "none"), (2, "transit", "none"),
        # Eski "transformer" degeri rol modeline donusur: tuketim noktasi.
        (3, "transit", "consumption"),
    ]
    assert plan.poles[1].name == "ara direk"
    assert plan.regions == {"B1": "B1"}


def test_quick_sheet_mevcut_hatta_SONUNA_ekler(db):
    """Hat Topoloji sayfasinda 2 direkle geliyorsa hizli satirlar 3'ten baslar."""
    data = _quick_sheet([
        ["B1", "H1", "39.5, 32.5", "", ""],
    ], with_topo=True)
    plan = g.parse_and_plan(data, db)
    assert [e.message for e in plan.errors] == []
    hizli = [p for p in plan.poles if p.sequence_no == 3]
    assert len(hizli) == 1 and hizli[0].latitude == 39.5


def test_quick_sheet_db_deki_hatta_ekler(db):
    """Hat DB'de 4 direkle varsa yeni direkler 5'ten devam eder."""
    r = Region(code="B1", name="Bolge 1"); db.add(r); db.flush()
    ln = Line(region_id=r.id, code="H1", name="Hat 1"); db.add(ln); db.flush()
    for i in range(1, 5):
        db.add(Pole(line_id=ln.id, sequence_no=i, latitude=39.0 + i * 0.01, longitude=32.0))
    db.flush()
    data = _quick_sheet([["B1", "H1", "39.9, 32.9", "", ""]])
    plan = g.parse_and_plan(data, db)
    assert plan.poles[0].sequence_no == 5


def test_wizard_plan_ve_apply(db):
    """Sihirbaz plani apply_plan'dan gecer: bolge+hat+direkler olusur,
    ters koordinat otomatik takaslanir."""
    plan = g.plan_for_wizard(
        db,
        region_code="B2", region_name="Bolge 2",
        line_code="H9", line_name="Hat 9",
        poles=[
            {"latitude": 39.0, "longitude": 32.0},
            {"latitude": 132.5, "longitude": 39.5},  # ters -> takas
            {"latitude": 39.2, "longitude": 32.2, "name": "son",
             "topology_role": "line_end", "energy_role": "consumption"},
        ],
    )
    assert [e.message for e in plan.errors] == []
    result = g.apply_plan(plan, db)
    assert result.regions_created == 1
    assert result.lines_created == 1
    assert result.poles_created == 3
    poles = db.query(Pole).order_by(Pole.sequence_no).all()
    assert poles[1].latitude == 39.5 and poles[1].longitude == 132.5
    assert poles[2].topology_role == "line_end"
    assert poles[2].energy_role == "consumption"


def test_template_hizli_yapistir_sayfasi_var(db):
    buf = g.build_template_workbook(db)
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    assert g.QUICK_SHEET in wb.sheetnames
    ws = wb[g.QUICK_SHEET]
    basliklar = [c.value for c in ws[1][: len(g.QUICK_COLUMNS)]]
    assert basliklar == g.QUICK_COLUMNS


def test_template_hizli_sayfa_ILK_ve_aktif(db):
    """Sablon hizli sayfayla ACILIR — kullanici karmasik Topoloji sayfasina
    dusup hizli yolu kacirmasin."""
    buf = g.build_template_workbook(db)
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    assert wb.sheetnames[0] == g.QUICK_SHEET
    assert wb.active.title == g.QUICK_SHEET


def test_quick_sheet_ornek_blok_forward_fill_KIRLETMEZ(db):
    """Ornek bloktaki ANKARA/TR-3, kullanicinin kendi satirlarina sizmaz."""
    data = _quick_sheet([
        ["ANKARA", "TR-3", "ornek: 39.92, 32.85", "d1", ""],
        ["", "", "ornek: 39.93, 32.86", "d2", ""],
        ["B-GERCEK", "H-GERCEK", "40.0, 33.0", "", ""],
        ["", "", "40.1, 33.1", "", ""],
    ])
    plan = g.parse_and_plan(data, db)
    assert [e.message for e in plan.errors] == []
    assert "ANKARA" not in plan.regions
    assert plan.regions == {"B-GERCEK": "B-GERCEK"}
    assert len(plan.poles) == 2


def test_kod_uret_deterministik():
    assert g._kod_uret("Merkez TR-3 Hattı") == "MERKEZ-TR-3-HATTI"
    assert g._kod_uret("Merkez TR-3 Hattı") == g._kod_uret("Merkez TR-3 Hattı")
    assert g._kod_uret("çĞüŞİö") == "CGUSIO"


def test_parse_yalniz_AD_ile_yeni_hat(db):
    """Kod kolonlari gizli: kullanici yalniz Bolge_Adi/Hat_Adi yazar, kod
    addan uretilir ve direkler o hatta baglanir."""
    data = _sheet([
        ["", "Merkez", "", "Ana Hat", 1, "", 39.0, 32.0, "", "", "", "", ""],
        ["", "", "", "", 2, "", 39.1, 32.1, "", "", "", "", ""],
    ])
    plan = g.parse_and_plan(data, db)
    assert [e.message for e in plan.errors] == []
    assert plan.regions == {"MERKEZ": "Merkez"}
    assert list(plan.lines.values())[0].code == "ANA-HAT"
    assert len(plan.poles) == 2


def test_template_kod_kolonlari_gizli_ve_cihaz_koordinati_dolu(db):
    """Sablonda Bolge_Kodu/Hat_Kodu kolonlari gizli; cihaz ara-satirinda
    otomatik enterpolasyonlu koordinat yazili (parser onu OKUMAZ)."""
    r = Region(code="B1", name="Bolge 1"); db.add(r); db.flush()
    ln = Line(region_id=r.id, code="H1", name="Hat 1"); db.add(ln); db.flush()
    p1 = Pole(line_id=ln.id, sequence_no=1, latitude=39.0, longitude=32.0)
    p2 = Pole(line_id=ln.id, sequence_no=2, latitude=40.0, longitude=33.0)
    db.add_all([p1, p2]); db.flush()
    d = Device(code="CIHAZ-1", name="c1", ip_address="10.0.0.9", latitude=0, longitude=0)
    db.add(d); db.flush()
    db.add(LineSegment(line_id=ln.id, from_pole_id=p1.id, to_pole_id=p2.id,
                       device_id=d.id, device_position_t=0.5)); db.flush()

    buf = g.build_template_workbook(db)
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    ws = wb["Topoloji"]
    assert ws.column_dimensions[g._col("Bolge_Kodu")].hidden is True
    assert ws.column_dimensions[g._col("Hat_Kodu")].hidden is True
    # Cihaz ara-satiri: Direk_Sira bos + Cihaz dolu; Enlem orta nokta ~39.5.
    cihaz_satiri = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(g.COLUMNS, list(row) + [None] * (len(g.COLUMNS) - len(row))))
        if g._clean(rec["Cihaz"]) == "CIHAZ-1":
            cihaz_satiri = rec
            break
    assert cihaz_satiri is not None
    assert abs(float(cihaz_satiri["Enlem"]) - 39.5) < 1e-6
    assert abs(float(cihaz_satiri["Boylam"]) - 32.5) < 1e-6
    # Round-trip: ayni dosya parse edildiginde cihaz koordinati direk sanilmaz.
    buf2 = io.BytesIO(); wb.save(buf2)
    plan = g.parse_and_plan(buf2.getvalue(), db)
    assert [e.message for e in plan.errors] == []
    assert len(plan.poles) == 2
    assert sum(len(p.devices) for p in plan.poles) == 1
