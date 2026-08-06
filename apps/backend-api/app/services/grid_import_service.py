"""Hat topolojisi Excel sablon uretme + import (onizlemeli).

Kullanici bos bir .xlsx sablonu indirir (Bolge/Hat/Direk/koordinat/cihaz),
doldurur, geri yukler. Sistem once ONIZLEME (parse_and_plan — DB'ye yazmadan
dogrulama) sonra ONAY ile COMMIT (apply_plan — tek transaction) yapar.

Tasarim:
  - Sablon tek sayfa duz tablo: her satir bir DIREK. Ayni Hat_Kodu altinda
    Direk_Sira 1..N ile o hattin direkleri olusur. Koordinatlar Pole'a yazilir.
  - Cihaz hucresi Excel dropdown (data validation): mevcut ATANMAMIS cihaz
    kodlari. Bir cihaz tek segmente baglanir (DB unique) — sablonda atanmislar
    listede yok, ayni cihaz iki satirda secilirse import HATA verir.
  - Cihaz OLUSTURMAZ; sadece mevcut cihazi koda gore segmente baglar.
  - Ekle/birlestir: Bolge/Hat kodu varsa guncelle, yoksa yarat; baska veriye
    dokunma. Direkler hat icin (sira bazli) upsert edilir.

Bu modul HTTP'den bagimsizdir (test edilebilir saf fonksiyonlar).
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.device import Device
from app.models.grid_topology import Line, LineSegment, Pole, Region

# Sablon kolonlari (sira onemli). Cihaz_Sira YOK — cihaz sirasi ara-satir
# sirasindan otomatik turer. Kolon harfleri COLUMNS index'inden hesaplanir
# (asagida _col), elle harf gomulmez.
COLUMNS = [
    "Bolge_Kodu",
    "Bolge_Adi",
    "Hat_Kodu",
    "Hat_Adi",
    "Direk_Sira",
    "Direk_Adi",
    "Enlem",
    "Boylam",
    "Direk_Tipi",
    "Cihaz",
    "Yon",          # FCI yon oryantasyonu: yesil | kirmizi
    "Bransman_Hat_Kodu",
    "Bransman_Direk_Sira",
]


def _col(name: str) -> str:
    """COLUMNS'taki kolon adinin Excel harfini dondur (A, B, ...). Kolon sirasi
    degisirse tum adresler otomatik guncellenir — elle harf gomme."""
    from openpyxl.utils import get_column_letter

    return get_column_letter(COLUMNS.index(name) + 1)


POLE_TYPES = ["pole", "transformer", "breaker"]

# Hizli Yapistir sayfasi: direk sirasi ve segment OTOMATIK — kullanici yalnizca
# koordinat listesi yapistirir. Tek 'Koordinat' hucresi "enlem, boylam" alir
# (Google Maps'ten kopyalanan bicim); istenirse ad/tip eklenir.
QUICK_SHEET = "Hizli_Yapistir"
QUICK_COLUMNS = ["Bolge", "Hat", "Koordinat", "Direk_Adi", "Direk_Tipi"]


def _parse_coord_pair(text: str) -> tuple[float, float] | None:
    """Tek hucre koordinati coz: "39.92, 32.85" / "39.92 32.85" / "39,92;32,85".

    AKILLI DAVRANISLAR (kullanici ugrasmasin diye):
      - Ayirac: virgul, noktali virgul, bosluk, tab — hangisi varsa.
      - Ondalik virgul destegi: "39,92 32,85" (iki sayi, ikiser parca) da cozulur.
      - Enlem/boylam TERS yapistirilmissa (|enlem|>90 ama takasla gecerli)
        SESSIZCE duzeltilir — Google Earth bazi bicimlerde boylami one koyar.
    """
    s = (text or "").strip()
    if not s:
        return None
    # Once standart ayiraclarla dene.
    for sep in (",", ";", "\t", " "):
        if sep in s:
            parts = [p for p in s.replace("\t", " ").split(sep) if p.strip()]
            if len(parts) == 2:
                lat = _to_float(parts[0])
                lon = _to_float(parts[1])
                if lat is not None and lon is not None:
                    return _maybe_swap(lat, lon)
            if sep == "," and len(parts) == 4:
                # "39,92 32,85" ondalik virgul + bosluk: "39","92 32","85" gibi
                # bolunmus olabilir — bosluga gore yeniden dene.
                sol, sag = s.split(None, 1) if " " in s else (None, None)
                if sol and sag:
                    lat = _to_float(sol)
                    lon = _to_float(sag)
                    if lat is not None and lon is not None:
                        return _maybe_swap(lat, lon)
    return None


def _maybe_swap(lat: float, lon: float) -> tuple[float, float] | None:
    """Gecersiz enlem, takasla gecerliyse takasla; ikisi de gecersizse None."""
    def ok(la: float, lo: float) -> bool:
        return -90 <= la <= 90 and -180 <= lo <= 180

    if ok(lat, lon):
        return (lat, lon)
    if ok(lon, lat):
        return (lon, lat)
    return None

# Yon (Excel) -> device_orientation (DB) esleme.
YON_TO_ORIENTATION = {
    "yesil": "green_forward",
    "yeşil": "green_forward",
    "kirmizi": "red_forward",
    "kırmızı": "red_forward",
    "green": "green_forward",
    "red": "red_forward",
}
ORIENTATION_TO_YON = {
    "green_forward": "yesil",
    "red_forward": "kirmizi",
}
YON_CHOICES = ["yesil", "kirmizi"]


def _sanitize_range_token(code: str) -> str:
    """Hat kodunu Excel named-range parcasina cevir: harf/rakam/altcizgi, harfle
    baslar. 'F-1' -> 'F_1'. Excel INDIRECT tarafi ile BIREBIR ayni donusum
    (SUBSTITUTE zinciri: bosluk ve tire -> altcizgi). Diger ozel karakterler de
    altcizgi olur ama Excel formulu yalniz bosluk+tire cevirir; bu yuzden hat
    kodlarinda yalniz harf/rakam/bosluk/tire varsayilir (UI zaten kisitli)."""
    out = []
    for ch in code:
        out.append(ch if (ch.isalnum() or ch == "_") else "_")
    token = "".join(out)
    if not token or not token[0].isalpha():
        token = "L_" + token
    return token[:200]


def _range_name(code: str) -> str:
    """Hat koduna karsilik gelen named-range adi (HAT_<token>)."""
    return "HAT_" + _sanitize_range_token(code)


# --------------------------------------------------------------------------- #
# Sablon uretme
# --------------------------------------------------------------------------- #
def _assigned_device_slots(db: Session) -> dict[int, tuple[str, str, int]]:
    """Cihaz id -> (bolge_kodu, hat_kodu, from_direk_sira).

    Mevcut sablon round-trip'inde cihaz zaten ayni slota bagli olabilir; bu hata
    degil, idempotent guncellemedir. Baska slot ise yine import hatasi.
    """
    regions = {r.id: r for r in db.scalars(select(Region)).all()}
    lines = {l.id: l for l in db.scalars(select(Line)).all()}
    poles = {p.id: p for p in db.scalars(select(Pole)).all()}
    out: dict[int, tuple[str, str, int]] = {}
    for s in db.scalars(select(LineSegment).where(LineSegment.device_id.is_not(None))).all():
        if s.device_id is None:
            continue
        line = lines.get(s.line_id)
        pole = poles.get(s.from_pole_id)
        region = regions.get(line.region_id) if line else None
        if line is None or pole is None or region is None:
            continue
        out[s.device_id] = (region.code, line.code, pole.sequence_no)
    return out


def _hex(color: str | None, fallback: str) -> str:
    """HEX rengi openpyxl formatina (AARRGGBB / RRGGBB) cevir. '#' at."""
    if not color:
        return fallback
    c = color.lstrip("#").strip()
    if len(c) == 6:
        return c.upper()
    return fallback


def _load_topology_rows(db: Session) -> list[dict]:
    """Sistemdeki tum Region->Line->Pole->Segment(cihaz) hiyerarsisini satirlara
    ac. Iki satir tipi ('kind'):
      - 'pole'   : bir DIREK (koordinat/tip burada, Cihaz bos).
      - 'device' : bir CIHAZ ara-satiri, ait oldugu direkten HEMEN SONRA gelir.
                   Cihaz o direkten baslayan segmente baglanir (from=direk,
                   to=sonraki direk). Coklu cihaz = ardisik birden fazla 'device'.
    Bolge/Hat sirali. Cihaz sirasi ara-satir sirasiyla verilir (device_position_t)."""
    regions = {r.id: r for r in db.scalars(select(Region).order_by(Region.name)).all()}
    lines = list(db.scalars(select(Line).order_by(Line.region_id, Line.code)).all())
    poles = list(db.scalars(select(Pole).order_by(Pole.line_id, Pole.sequence_no)).all())
    segments = list(db.scalars(select(LineSegment)).all())
    devices = {d.id: d for d in db.scalars(select(Device)).all()}

    poles_by_line: dict[int, list[Pole]] = {}
    for p in poles:
        poles_by_line.setdefault(p.line_id, []).append(p)

    # (line_id, from_seq) -> [ (t, device_code, orientation) ]
    seg_by_slot: dict[tuple[int, int], list[tuple[float, str, str | None]]] = {}
    seq_by_pole = {p.id: p.sequence_no for p in poles}
    for s in segments:
        if s.device_id is None:
            continue
        dev = devices.get(s.device_id)
        if dev is None:
            continue
        from_seq = seq_by_pole.get(s.from_pole_id)
        if from_seq is None:
            continue
        t = s.device_position_t if s.device_position_t is not None else 0.5
        seg_by_slot.setdefault((s.line_id, from_seq), []).append(
            (t, dev.code, s.device_orientation)
        )
    # Her slottaki cihazlari t'ye gore sirala (ara-satir sirasi).
    for key in seg_by_slot:
        seg_by_slot[key].sort(key=lambda x: x[0])

    rows: list[dict] = []
    lines_by_region: dict[int, list[Line]] = {}
    for ln in lines:
        lines_by_region.setdefault(ln.region_id, []).append(ln)

    for region_id, region in regions.items():
        region_lines = lines_by_region.get(region_id, [])
        for ln in region_lines:
            line_poles = poles_by_line.get(ln.id, [])
            branch_line_code = ""
            branch_pole_seq = ""
            if ln.branched_from_pole_id is not None:
                bp_seq = seq_by_pole.get(ln.branched_from_pole_id)
                for other in lines:
                    if any(p.id == ln.branched_from_pole_id for p in poles_by_line.get(other.id, [])):
                        branch_line_code = other.code
                        break
                branch_pole_seq = bp_seq if bp_seq is not None else ""
            for p in line_poles:
                # Direk satiri (cihazsiz).
                rows.append({
                    "kind": "pole",
                    "region": region, "line": ln, "pole": p,
                    "branch_line_code": branch_line_code,
                    "branch_pole_seq": branch_pole_seq,
                })
                # O direkten baslayan slottaki cihazlar -> ayri ara-satirlar.
                slot = seg_by_slot.get((ln.id, p.sequence_no), [])
                for _, dev_code, orient in slot:
                    rows.append({
                        "kind": "device",
                        "region": region, "line": ln, "pole": p,
                        "device_code": dev_code,
                        "yon": ORIENTATION_TO_YON.get(orient or "", ""),
                    })
    return rows


def build_template_workbook(db: Session) -> io.BytesIO:
    """Import sablonu (.xlsx). MEVCUT topolojiyi DOLU getirir + agac gorunum
    (Bolge>Hat>Direk katlanabilir outline). Cihazlar direkler arasinda AYRI
    ara-satirda. Cihaz/Direk_Tipi/Yon dropdown; bransman bagimli dropdown;
    tekrar eden cihaz kirmizi (kosullu bicim).

    Bos satirlar altta yeni ekleme icin. Doner: BytesIO."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    devices = list(db.scalars(select(Device).order_by(Device.code)).all())
    all_device_codes = [d.code for d in devices]
    all_line_codes = list(db.scalars(select(Line.code).order_by(Line.code)).all())
    # Hat kodu -> o hattin direk sira listesi (bransman bagimli dropdown icin).
    line_pole_seqs: dict[str, list[int]] = {}
    for ln in db.scalars(select(Line)).all():
        seqs = sorted(
            p.sequence_no
            for p in db.scalars(select(Pole).where(Pole.line_id == ln.id)).all()
        )
        line_pole_seqs[ln.code] = seqs

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws = wb.active
    ws.title = "Topoloji"
    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Agac gorunum: Bolge (level 0 baslik) > Hat (level 1 baslik) >
    # Direk + Cihaz ara-satirlari (level 2). Bolge/Hat kolonlari yalniz grup
    # basinda dolu (parse forward-fill yapar). Cihaz ara-satirinda Direk_Sira
    # ve koordinat/tip BOS, Cihaz+Yon dolu, Direk_Adi = "-> N-M arasi" etiketi.
    topo_rows = _load_topology_rows(db)
    excel_row = 2
    last_region_code = None
    last_line_key = None
    region_fill_cache: dict[str, PatternFill] = {}
    line_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    device_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Sonraki direk sirasini bulmak icin: ayni hattaki bir sonraki 'pole' satiri.
    def _next_pole_seq(idx: int, line_id: int) -> int | None:
        for j in range(idx + 1, len(topo_rows)):
            nr = topo_rows[j]
            if nr["line"].id != line_id:
                return None
            if nr["kind"] == "pole":
                return nr["pole"].sequence_no
        return None

    for i, r in enumerate(topo_rows):
        region = r["region"]
        line = r["line"]
        pole = r["pole"]
        region_code = region.code
        line_key = f"{region_code}|{line.code}"
        is_new_region = region_code != last_region_code
        is_new_line = line_key != last_line_key

        if r["kind"] == "pole":
            vals = [
                region_code if is_new_region else "",
                region.name if is_new_region else "",
                line.code if is_new_line else "",
                line.name if is_new_line else "",
                pole.sequence_no,
                pole.name or "",
                pole.latitude,
                pole.longitude,
                pole.pole_type or "pole",
                "", "",  # Cihaz, Yon bos
                r["branch_line_code"] if is_new_line else "",
                r["branch_pole_seq"] if is_new_line else "",
            ]
        else:  # device ara-satiri
            nxt = _next_pole_seq(i, line.id)
            label = f"-> {pole.sequence_no}-{nxt} arasi" if nxt else f"-> {pole.sequence_no} sonrasi"
            vals = [
                "", "", "", "",   # bolge/hat bos (ust satirdan gelir)
                "",               # Direk_Sira BOS -> cihaz satiri isareti
                label,            # Direk_Adi = bilgilendirici etiket
                "", "", "",       # Enlem/Boylam/Direk_Tipi bos
                r["device_code"],
                r["yon"],
                "", "",           # bransman bos
            ]
        ws.append(vals)

        # Outline seviyeleri.
        if is_new_line:
            ws.row_dimensions[excel_row].outline_level = 1  # hat basligi
        else:
            ws.row_dimensions[excel_row].outline_level = 2  # direk/cihaz

        # Renklendirme.
        if is_new_region:
            rc = _hex(region.color, "DBEAFE")
            if rc not in region_fill_cache:
                region_fill_cache[rc] = PatternFill(start_color=rc, end_color=rc, fill_type="solid")
            fill = region_fill_cache[rc]
            for col in (1, 2):
                ws.cell(row=excel_row, column=col).fill = fill
                ws.cell(row=excel_row, column=col).font = Font(bold=True)
        if is_new_line:
            for col in (3, 4):
                ws.cell(row=excel_row, column=col).fill = line_fill
                ws.cell(row=excel_row, column=col).font = Font(bold=True)
        if r["kind"] == "device":
            # Cihaz ara-satirini hafifce boya (gorsel ayrim).
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=excel_row, column=col).fill = device_fill

        last_region_code = region_code
        last_line_key = line_key
        excel_row += 1

    ws.sheet_properties.outlinePr.summaryBelow = False

    # Bos ekleme satirlari (yeni hat/direk/cihaz icin).
    blank_end = excel_row + 200

    widths = [12, 20, 10, 18, 10, 18, 12, 12, 13, 18, 10, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[_col(COLUMNS[i - 1])].width = w
    ws.freeze_panes = "A2"

    _add_dropdowns(
        wb, ws, all_device_codes, all_line_codes, line_pole_seqs,
        last_data_row=blank_end,
    )
    _add_conditional_formats(ws, last_data_row=blank_end)

    # --- Hizli Yapistir sheet (toplu koordinat girisi) ---
    _add_quick_sheet(wb, header_font, header_fill)

    # --- Yardim sheet ---
    _add_help_sheet(wb, header_font, header_fill)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _add_quick_sheet(wb, header_font, header_fill) -> None:
    """Hizli Yapistir: direkleri TEK SEFERDE yapistirmak icin sade sayfa.

    Kullanim: Bolge + Hat bir kez yazilir (alt satirlarda bos = ustteki),
    Koordinat kolonuna "enlem, boylam" listesi komple yapistirilir. Sira
    numarasi ve segmentler OTOMATIK; hat mevcutsa direkler SONUNA eklenir.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet(QUICK_SHEET)
    ws.append(QUICK_COLUMNS)
    for col_idx in range(1, len(QUICK_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [16, 16, 30, 18, 13]
    for i, w in enumerate(widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Direk_Tipi dropdown (E kolonu).
    dv = DataValidation(
        type="list", formula1='"' + ",".join(POLE_TYPES) + '"',
        allow_blank=True, showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add("E2:E5001")

    # Sag tarafta kisa kullanim notu (parser ilk 5 kolonu okur; G serbest).
    notlar = [
        "NASIL KULLANILIR",
        "1) Bolge ve Hat adini ILK satira bir kez yazin (altta bos birakin).",
        "2) Koordinat kolonuna 'enlem, boylam' listesini KOMPLE yapistirin.",
        "   Google Maps kopyasi (39.92042, 32.85411) dogrudan calisir.",
        "   'enlem boylam' (bosluklu) ve ondalik virgul de kabul edilir.",
        "3) Sira ve segmentler OTOMATIK: satir sirasi = direk sirasi.",
        "   Hat zaten varsa yeni direkler hattin SONUNA eklenir.",
        "4) Direk_Adi ve Direk_Tipi istege bagli (bos = pole).",
        "Ayni dosyada birden fazla hat: yeni hatta gecerken Bolge/Hat yazin.",
    ]
    note_font = Font(color="475569", size=10)
    for i, satir in enumerate(notlar):
        c = ws.cell(row=2 + i, column=7, value=satir)
        c.font = Font(bold=True, size=11) if i == 0 else note_font
    ws.column_dimensions["G"].width = 70

    # Ornek iki satir (acik gri) — parser icin zararsiz ORNEK etiketli Bolge
    # kullanilmaz: ornek koordinatlar bos birakilir, sadece bicim gosterilir.
    ornek_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    ws.cell(row=2, column=3, value="ornek: 39.92042, 32.85411").fill = ornek_fill
    ws.cell(row=2, column=3).font = Font(italic=True, color="94A3B8")


def _add_conditional_formats(ws, last_data_row: int) -> None:
    """Tekrar eden Cihaz kodunu KIRMIZI boya (kosullu bicim). Ayni cihaz iki
    satirda secilirse kullanici aninda gorur; import de hata verir."""
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle

    cihaz = _col("Cihaz")
    end = max(last_data_row, 2)
    rng = f"{cihaz}2:{cihaz}{end}"
    red = DifferentialStyle(
        fill=PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        font=Font(color="B91C1C", bold=True),
    )
    # Bos degilse VE ayni deger 1'den fazla ise -> kirmizi.
    formula = f'AND({cihaz}2<>"",COUNTIF(${cihaz}$2:${cihaz}${end},{cihaz}2)>1)'
    rule = Rule(type="expression", formula=[formula], dxf=red, stopIfTrue=False)
    ws.conditional_formatting.add(rng, rule)


def _add_dropdowns(
    wb, ws, all_device_codes, all_line_codes, line_pole_seqs, last_data_row: int
) -> None:
    """Acilir listeler: Direk_Tipi, Cihaz, Yon (duz) + Bransman_Hat_Kodu (duz,
    tum hatlar) + Bransman_Direk_Sira (BAGIMLI — secili hattin direk siralari,
    INDIRECT ile). Kolon harfleri COLUMNS index'inden turer (elle harf yok)."""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import quote_sheetname

    end = max(last_data_row, 2)
    c_tip = _col("Direk_Tipi")
    c_cihaz = _col("Cihaz")
    c_yon = _col("Yon")
    c_br_hat = _col("Bransman_Hat_Kodu")
    c_br_dir = _col("Bransman_Direk_Sira")

    def _add_list_dv(col: str, formula1: str) -> None:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{end}")

    # Direk_Tipi + Yon: inline liste.
    _add_list_dv(c_tip, '"' + ",".join(POLE_TYPES) + '"')
    _add_list_dv(c_yon, '"' + ",".join(YON_CHOICES) + '"')

    # Cihaz: gizli "Cihazlar" sheet'inden (atanmislar dahil — dolu satirlar
    # kendi cihazini gostersin).
    ws_dev = wb.create_sheet("Cihazlar")
    ws_dev.append(["Cihaz_Kodu"])
    ws_dev.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    ws_dev.cell(row=1, column=1).fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    for code in all_device_codes:
        ws_dev.append([code])
    ws_dev.column_dimensions["A"].width = 24
    ws_dev.sheet_state = "hidden"
    if all_device_codes:
        last = len(all_device_codes) + 1
        _add_list_dv(c_cihaz, f"Cihazlar!$A$2:$A${last}")

    # Bransman_Hat_Kodu: gizli "Hatlar" sheet A kolonundan (tum hat kodlari).
    # Her hat icin B..N kolonlarina direk siralari yazilir + named range HAT_<token>.
    ws_ln = wb.create_sheet("Hatlar")
    ws_ln.append(["Hat_Kodu", "Direk_Siralari"])
    ws_ln.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    ws_ln.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")
    for c in (1, 2):
        ws_ln.cell(row=1, column=c).fill = PatternFill(
            start_color="1F2937", end_color="1F2937", fill_type="solid"
        )
    ws_ln.column_dimensions["A"].width = 18
    ws_ln.sheet_state = "hidden"

    quoted = quote_sheetname("Hatlar")
    used_tokens: set[str] = set()
    row = 2
    for code in all_line_codes:
        seqs = line_pole_seqs.get(code, [])
        ws_ln.cell(row=row, column=1, value=code)
        # Direk siralarini C..N sutunlarina yatay yaz (named range yatay aralik).
        for j, seq in enumerate(seqs):
            ws_ln.cell(row=row, column=3 + j, value=seq)
        # Named range: HAT_<token>. Cakisma olursa suffix.
        base = _range_name(code)
        token = base
        n = 2
        while token in used_tokens:
            token = f"{base}_{n}"
            n += 1
        used_tokens.add(token)
        if seqs:
            from openpyxl.utils import get_column_letter as _gcl
            first = _gcl(3)
            lastc = _gcl(3 + len(seqs) - 1)
            ref = f"{quoted}!${first}${row}:${lastc}${row}"
            wb.defined_names.add(DefinedName(token, attr_text=ref))
        row += 1

    if all_line_codes:
        last_ln = len(all_line_codes) + 1
        _add_list_dv(c_br_hat, f"Hatlar!$A$2:$A${last_ln}")

        # Bagimli dropdown: secili hat koduna gore direk siralari. INDIRECT +
        # SUBSTITUTE zinciri _sanitize_range_token ile BIREBIR ayni (bosluk+tire
        # -> altcizgi). Kod cakismasinda suffix'li named range INDIRECT'te
        # bulunmaz (nadir; o durumda dropdown bos gelir ama import yine dogrular).
        indirect = (
            f'INDIRECT("HAT_"&SUBSTITUTE(SUBSTITUTE({c_br_hat}2," ","_"),"-","_"))'
        )
        _add_list_dv(c_br_dir, indirect)


def _add_help_sheet(wb, header_font, header_fill) -> None:
    ws_help = wb.create_sheet("Yardim")
    help_rows = [
        ["Kolon / Konu", "Açıklama"],
        ["YAPISI", "İki satır tipi vardır: DİREK satırı (Direk_Sira dolu) ve CİHAZ ara-satırı (Direk_Sira BOŞ, Cihaz dolu). Cihaz ara-satırı, hemen ÜSTÜNDEKİ direkten sonraki segmente bağlanır."],
        ["Bolge_Kodu / Bolge_Adi", "Bölge kodu ve adı. Grup başında bir kez yazılır; alt satırlarda boş bırakılabilir (üstteki geçerli sayılır)."],
        ["Hat_Kodu / Hat_Adi", "Hat kodu ve adı. Aynı bölge içinde kod benzersiz. Grup başında bir kez yazılır."],
        ["Direk_Sira", "DİREK satırında 1'den başlayan sıra. Boşluksuz artmalı. Cihaz ara-satırında BOŞ bırakılır."],
        ["Direk_Adi", "Direk satırında opsiyonel ad. Cihaz ara-satırında '-> N-M arası' bilgi etiketi (elle doldurmaya gerek yok)."],
        ["Enlem / Boylam", "Direk satırında ondalık derece. Enlem -90..90, Boylam -180..180."],
        ["Direk_Tipi", "pole / transformer / breaker. Boş = pole."],
        ["Cihaz", "Bir CİHAZ ara-satırına yazılır: üstteki direkten sonraki segmente bağlanacak cihaz (açılır liste). Aynı iki direk arasında birden fazla cihaz için üst üste birden çok cihaz ara-satırı ekleyin; sıra satır sırasıdır (otomatik)."],
        ["Yon", "Cihazın FCI yön oryantasyonu: yesil (A ucu ileri) / kirmizi (B ucu ileri). Opsiyonel."],
        ["Bransman_Hat_Kodu", "Bu hat başka hattın direğinden dallanıyorsa o hattın kodu (açılır liste). Hat başı satırına yazılır."],
        ["Bransman_Direk_Sira", "Bağlanılacak direk sırası. Bransman_Hat_Kodu seçilince açılır liste O HATTIN direk sıralarını gösterir (bağımlı liste)."],
        ["Tekrar eden cihaz", "Aynı cihaz iki kez seçilirse hücre KIRMIZI boyanır ve içe aktarımda hata verir; bir cihaz yalnız bir segmente bağlanır."],
    ]
    for r in help_rows:
        ws_help.append(r)
    ws_help.cell(row=1, column=1).font = header_font
    ws_help.cell(row=1, column=2).font = header_font
    ws_help.cell(row=1, column=1).fill = header_fill
    ws_help.cell(row=1, column=2).fill = header_fill
    ws_help.column_dimensions["A"].width = 28
    ws_help.column_dimensions["B"].width = 90


# --------------------------------------------------------------------------- #
# Parse + plan (onizleme icin — DB'ye yazmaz)
# --------------------------------------------------------------------------- #
@dataclass
class RowError:
    row: int          # 1-indexli Excel satir no (baslik = 1, veri 2'den)
    message: str


@dataclass
class ParsedDevice:
    """Bir slota (direk-sonrasi segment) baglanacak cihaz + sira + yon."""
    device_code: str
    cihaz_sira: int          # slot icindeki 1-tabanli sira
    orientation: str | None  # device_orientation (green_forward/red_forward/None)
    excel_row: int


@dataclass
class ParsedPole:
    region_code: str
    line_code: str
    sequence_no: int
    name: str | None
    latitude: float
    longitude: float
    pole_type: str
    excel_row: int
    # Bu direkten baslayan segmentteki cihaz(lar). Coklu olabilir (slot ici sira).
    devices: list[ParsedDevice] = field(default_factory=list)


@dataclass
class ParsedLine:
    region_code: str
    code: str
    name: str
    branch_line_code: str | None = None
    branch_pole_seq: int | None = None


@dataclass
class ImportPlan:
    regions: dict[str, str] = field(default_factory=dict)         # code -> name
    lines: dict[str, ParsedLine] = field(default_factory=dict)    # "region|line" -> ParsedLine
    poles: list[ParsedPole] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    def counts(self) -> dict[str, int]:
        device_count = sum(len(p.devices) for p in self.poles)
        return {
            "regions": len(self.regions),
            "lines": len(self.lines),
            "poles": len(self.poles),
            "devices": device_count,
            "errors": len(self.errors),
        }


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        # Excel virgullu ondalik gelebilir.
        if isinstance(v, str):
            v = v.strip().replace(",", ".")
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        if isinstance(v, float):
            return int(v)
        return int(str(v).strip())
    except (ValueError, TypeError):
        return None


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_and_plan(file_bytes: bytes, db: Session) -> ImportPlan:
    """Excel'i oku, dogrula, DB'ye YAZMADAN plan cikar. Satir-bazli hatalar
    plan.errors'da toplanir; hatasiz satirlar yine plana girer (commit onlari
    uygular, hataliyi atlar)."""
    from openpyxl import load_workbook

    plan = ImportPlan()

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        plan.errors.append(RowError(0, "Dosya okunamadı. Geçerli bir .xlsx dosyası mı?"))
        return plan

    if "Topoloji" in wb.sheetnames:
        ws = wb["Topoloji"]
    else:
        ws = wb.active

    # Mevcut cihazlar (kod -> id) ve atanmis cihazlar.
    device_by_code = {d.code: d for d in db.scalars(select(Device)).all()}
    assigned_slots = _assigned_device_slots(db)
    used_device_codes: set[str] = set()   # bu import icinde kullanilanlar (duplicate yakala)
    # (region_code, line_code, seq) -> ParsedPole (ayni direk coklu cihazda tekrar satirda gelir)
    pole_by_key: dict[tuple[str, str, int], ParsedPole] = {}
    # (region_code, line_code, seq) -> kullanilan cihaz_sira set (slot ici sira cakismasi yakala)
    slot_orders: dict[tuple[str, str, int], set[int]] = {}

    # Forward-fill: Bolge/Hat kolonlari agac gorunumunde grup basinda dolu,
    # alt satirlarda bos. Bos gelirse ustteki gecerli sayilir.
    ff_region_code = ""
    ff_region_name = ""
    ff_hat_code = ""
    ff_hat_name = ""
    ff_branch_line = ""
    ff_branch_seq: int | None = None
    # Cihaz ara-satiri en son goruilen DIREGE baglanir.
    last_pole_key: tuple[str, str, int] | None = None

    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    for idx, raw in enumerate(rows_iter, start=2):
        # Bos satiri atla.
        if raw is None or all((c is None or _clean(c) == "") for c in raw):
            continue
        cells = list(raw) + [None] * (len(COLUMNS) - len(raw))
        rec = dict(zip(COLUMNS, cells))

        # Forward-fill bolge/hat.
        rc = _clean(rec["Bolge_Kodu"])
        if rc:
            ff_region_code = rc
            ff_region_name = _clean(rec["Bolge_Adi"]) or rc
        hc = _clean(rec["Hat_Kodu"])
        if hc:
            ff_hat_code = hc
            ff_hat_name = _clean(rec["Hat_Adi"]) or hc
            # Bransman bilgisi hat basinda tasinir.
            ff_branch_line = _clean(rec["Bransman_Hat_Kodu"])
            ff_branch_seq = _to_int(rec["Bransman_Direk_Sira"])
            last_pole_key = None  # yeni hat -> cihaz baglama sifirla

        region_code = ff_region_code
        hat_code = ff_hat_code
        seq = _to_int(rec["Direk_Sira"])
        device_code_raw = _clean(rec["Cihaz"])

        # ---- CIHAZ ARA-SATIRI: Direk_Sira bos + Cihaz dolu ----
        if seq is None and device_code_raw:
            _parse_device_row(
                idx, rec, device_code_raw, last_pole_key, plan,
                device_by_code, assigned_slots, used_device_codes,
                pole_by_key, slot_orders,
            )
            continue

        # ---- DIREK SATIRI: Direk_Sira dolu ----
        if seq is None:
            # Ne direk ne cihaz — atla (bos/etiket satiri).
            continue

        lat = _to_float(rec["Enlem"])
        lon = _to_float(rec["Boylam"])

        if not region_code:
            plan.errors.append(RowError(idx, "Bolge_Kodu boş (üstte de yok)."))
            continue
        if not hat_code:
            plan.errors.append(RowError(idx, "Hat_Kodu boş (üstte de yok)."))
            continue
        if seq < 1:
            plan.errors.append(RowError(idx, "Direk_Sira geçersiz (1'den başlamalı)."))
            continue

        pole_key = (region_code, hat_code, seq)
        existing_pole = pole_by_key.get(pole_key)

        # Direk ilk kez goruluyor -> koordinat + tip dogrula, olustur.
        if existing_pole is None:
            if lat is None or not (-90 <= lat <= 90):
                plan.errors.append(RowError(idx, f"Enlem geçersiz: {rec['Enlem']!r} (-90..90)."))
                continue
            if lon is None or not (-180 <= lon <= 180):
                plan.errors.append(RowError(idx, f"Boylam geçersiz: {rec['Boylam']!r} (-180..180)."))
                continue
            pole_type = _clean(rec["Direk_Tipi"]).lower() or "pole"
            if pole_type not in POLE_TYPES:
                plan.errors.append(RowError(idx, f"Direk_Tipi geçersiz: {pole_type!r}."))
                continue
            # Bolge + hat plana ekle (upsert — kod bazli).
            if region_code not in plan.regions:
                plan.regions[region_code] = ff_region_name
            line_key = f"{region_code}|{hat_code}"
            if line_key not in plan.lines:
                plan.lines[line_key] = ParsedLine(
                    region_code=region_code, code=hat_code, name=ff_hat_name or hat_code,
                    branch_line_code=ff_branch_line or None, branch_pole_seq=ff_branch_seq,
                )
            existing_pole = ParsedPole(
                region_code=region_code, line_code=hat_code, sequence_no=seq,
                name=_clean(rec["Direk_Adi"]) or None,
                latitude=lat, longitude=lon, pole_type=pole_type, excel_row=idx,
            )
            pole_by_key[pole_key] = existing_pole
            plan.poles.append(existing_pole)
            slot_orders[pole_key] = set()

        last_pole_key = pole_key

        # Geriye donuk uyum: cihaz AYNI direk satirinda da yazilmis olabilir
        # (eski sablon veya kullanici elle). Varsa yine bagla.
        if device_code_raw:
            _parse_device_row(
                idx, rec, device_code_raw, pole_key, plan,
                device_by_code, assigned_slots, used_device_codes,
                pole_by_key, slot_orders,
            )

    # ---- HIZLI YAPISTIR sayfasi (varsa) ----
    if QUICK_SHEET in wb.sheetnames:
        _parse_quick_sheet(wb[QUICK_SHEET], plan, db, pole_by_key, slot_orders)

    wb.close()
    return plan


def _next_seq_start(
    db: Session, region_code: str, line_code: str,
    pole_by_key: dict[tuple[str, str, int], "ParsedPole"],
) -> int:
    """Hatta EKLENECEK ilk sira: DB'deki + plandaki (Topoloji sayfasi) en
    buyuk siranin devami. Hat yoksa 1."""
    en_buyuk = 0
    region = db.scalar(select(Region).where(Region.code == region_code))
    if region is not None:
        line = db.scalar(
            select(Line).where(Line.region_id == region.id, Line.code == line_code)
        )
        if line is not None:
            db_max = db.scalar(
                select(func.max(Pole.sequence_no)).where(Pole.line_id == line.id)
            )
            en_buyuk = int(db_max or 0)
    for (rc, lc, seq) in pole_by_key:
        if rc == region_code and lc == line_code and seq > en_buyuk:
            en_buyuk = seq
    return en_buyuk + 1


def _parse_quick_sheet(
    ws, plan: ImportPlan, db: Session,
    pole_by_key: dict[tuple[str, str, int], "ParsedPole"],
    slot_orders: dict[tuple[str, str, int], set[int]],
) -> None:
    """Hizli Yapistir satirlarini plana cevir.

    Kurallar:
      - Bolge/Hat forward-fill (bir kez yazilir). Tek hucre hem kod hem ad
        sayilir — hizli giriste ayri kod/ad ayrimi kullaniciyi ugrastirir.
      - Koordinat tek hucrede "enlem, boylam"; ters yapistirma otomatik takas.
      - Sira OTOMATIK: satir sirasi = direk sirasi; hat mevcutsa (DB'de ya da
        Topoloji sayfasinda) yeni direkler SONUNA eklenir.
      - "ornek" ile baslayan koordinat hucreleri yok sayilir (sablon ornegi).
    """
    ff_bolge = ""
    ff_hat = ""
    seq_next: dict[tuple[str, str], int] = {}

    for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if raw is None or all((c is None or _clean(c) == "") for c in raw):
            continue
        cells = list(raw) + [None] * (len(QUICK_COLUMNS) - len(raw))
        rec = dict(zip(QUICK_COLUMNS, cells[: len(QUICK_COLUMNS)]))

        b = _clean(rec["Bolge"])
        if b:
            ff_bolge = b
        h = _clean(rec["Hat"])
        if h:
            ff_hat = h

        koor = _clean(rec["Koordinat"])
        if not koor or koor.lower().startswith("ornek"):
            continue
        if not ff_bolge:
            plan.errors.append(RowError(idx, f"{QUICK_SHEET}: Bolge boş (üstte de yok)."))
            continue
        if not ff_hat:
            plan.errors.append(RowError(idx, f"{QUICK_SHEET}: Hat boş (üstte de yok)."))
            continue

        cift = _parse_coord_pair(koor)
        if cift is None:
            plan.errors.append(
                RowError(idx, f"{QUICK_SHEET}: Koordinat çözülemedi: {koor!r} (beklenen: 'enlem, boylam').")
            )
            continue
        lat, lon = cift

        pole_type = _clean(rec["Direk_Tipi"]).lower() or "pole"
        if pole_type not in POLE_TYPES:
            plan.errors.append(RowError(idx, f"{QUICK_SHEET}: Direk_Tipi geçersiz: {pole_type!r}."))
            continue

        anahtar = (ff_bolge, ff_hat)
        if anahtar not in seq_next:
            seq_next[anahtar] = _next_seq_start(db, ff_bolge, ff_hat, pole_by_key)
        seq = seq_next[anahtar]
        seq_next[anahtar] = seq + 1

        if ff_bolge not in plan.regions:
            plan.regions[ff_bolge] = ff_bolge
        line_key = f"{ff_bolge}|{ff_hat}"
        if line_key not in plan.lines:
            plan.lines[line_key] = ParsedLine(
                region_code=ff_bolge, code=ff_hat, name=ff_hat,
            )
        pp = ParsedPole(
            region_code=ff_bolge, line_code=ff_hat, sequence_no=seq,
            name=_clean(rec["Direk_Adi"]) or None,
            latitude=lat, longitude=lon, pole_type=pole_type, excel_row=idx,
        )
        pole_by_key[(ff_bolge, ff_hat, seq)] = pp
        plan.poles.append(pp)
        slot_orders[(ff_bolge, ff_hat, seq)] = set()


# --------------------------------------------------------------------------- #
# Soru-cevap sihirbazi — tek hatti JSON'dan plana cevir
# --------------------------------------------------------------------------- #
def plan_for_wizard(
    db: Session,
    *,
    region_code: str,
    region_name: str | None,
    line_code: str,
    line_name: str | None,
    poles: list[dict],
    branch_line_code: str | None = None,
    branch_pole_seq: int | None = None,
) -> ImportPlan:
    """Sihirbaz girdisini (tek hat + koordinat listesi) ImportPlan'a cevir.

    Excel yoluyla AYNI apply_plan'dan gecer — tek dogruluk kaynagi. Hat
    mevcutsa direkler sonuna eklenir (sira otomatik devam eder).
    """
    plan = ImportPlan()
    region_code = region_code.strip()
    line_code = line_code.strip()
    if not region_code:
        plan.errors.append(RowError(0, "Bölge boş."))
        return plan
    if not line_code:
        plan.errors.append(RowError(0, "Hat kodu boş."))
        return plan

    plan.regions[region_code] = (region_name or region_code).strip() or region_code
    plan.lines[f"{region_code}|{line_code}"] = ParsedLine(
        region_code=region_code, code=line_code,
        name=(line_name or line_code).strip() or line_code,
        branch_line_code=(branch_line_code or "").strip() or None,
        branch_pole_seq=branch_pole_seq,
    )

    bos: dict[tuple[str, str, int], ParsedPole] = {}
    seq = _next_seq_start(db, region_code, line_code, bos)
    for i, p in enumerate(poles, start=1):
        lat = _to_float(p.get("latitude"))
        lon = _to_float(p.get("longitude"))
        cift = _maybe_swap(lat, lon) if lat is not None and lon is not None else None
        if cift is None:
            plan.errors.append(RowError(i, f"{i}. direk koordinatı geçersiz."))
            continue
        pole_type = _clean(p.get("pole_type")).lower() or "pole"
        if pole_type not in POLE_TYPES:
            plan.errors.append(RowError(i, f"{i}. direk tipi geçersiz: {pole_type!r}."))
            continue
        plan.poles.append(
            ParsedPole(
                region_code=region_code, line_code=line_code, sequence_no=seq,
                name=_clean(p.get("name")) or None,
                latitude=cift[0], longitude=cift[1],
                pole_type=pole_type, excel_row=i,
            )
        )
        seq += 1
    return plan


def _parse_device_row(
    idx, rec, device_code_raw, pole_key, plan,
    device_by_code, assigned_slots, used_device_codes,
    pole_by_key, slot_orders,
) -> None:
    """Bir cihazi (ara-satir veya direk satirindaki Cihaz hucresi) ait oldugu
    direk slotuna ekle. Cihaz_Sira YOK — sira ara-satir sirasidir (append)."""
    if pole_key is None:
        plan.errors.append(RowError(idx, f"Cihaz {device_code_raw!r} bir direğe bağlı değil (üstte direk yok)."))
        return
    existing_pole = pole_by_key.get(pole_key)
    if existing_pole is None:
        plan.errors.append(RowError(idx, f"Cihaz {device_code_raw!r} için geçerli direk yok."))
        return
    dev = device_by_code.get(device_code_raw)
    if dev is None:
        plan.errors.append(RowError(idx, f"Cihaz bulunamadı: {device_code_raw!r}."))
        return
    if device_code_raw in used_device_codes:
        plan.errors.append(RowError(idx, f"Cihaz {device_code_raw!r} birden fazla kez seçilmiş."))
        return
    region_code, hat_code, seq = pole_key
    assigned_slot = assigned_slots.get(dev.id)
    if assigned_slot is not None and assigned_slot != (region_code, hat_code, seq):
        plan.errors.append(RowError(idx, f"Cihaz {device_code_raw!r} zaten başka segmente bağlı."))
        return

    # Yon (opsiyonel).
    yon_raw = _clean(rec["Yon"]).lower()
    orientation = YON_TO_ORIENTATION.get(yon_raw) if yon_raw else None
    if yon_raw and orientation is None:
        plan.errors.append(RowError(idx, f"Yon geçersiz: {yon_raw!r} (yesil/kirmizi)."))
        return

    # Sira = slottaki append sirasi (otomatik).
    cihaz_sira = len(slot_orders.setdefault(pole_key, set())) + 1
    slot_orders[pole_key].add(cihaz_sira)
    used_device_codes.add(device_code_raw)
    existing_pole.devices.append(
        ParsedDevice(
            device_code=device_code_raw, cihaz_sira=cihaz_sira,
            orientation=orientation, excel_row=idx,
        )
    )


# --------------------------------------------------------------------------- #
# Apply (commit — tek transaction)
# --------------------------------------------------------------------------- #
@dataclass
class ImportResult:
    regions_created: int = 0
    regions_updated: int = 0
    lines_created: int = 0
    lines_updated: int = 0
    poles_created: int = 0
    poles_updated: int = 0
    segments_created: int = 0
    skipped: int = 0
    errors: list[RowError] = field(default_factory=list)


def apply_plan(plan: ImportPlan, db: Session) -> ImportResult:
    """Plani tek transaction'da uygula. Sira: Region -> Line -> Pole ->
    Segment (+ bransman 2. gecis). Caller commit eder (record_event sonrasi).

    Coklu cihaz slotlarinda device_position_t Excel Cihaz_Sira'ya gore yazilir;
    tek cihaz 0.5 alir. Ekstra konum senkronuna gerek yok."""
    result = ImportResult(errors=list(plan.errors))

    # 1) Region upsert (kod bazli).
    region_by_code: dict[str, Region] = {}
    for code, name in plan.regions.items():
        existing = db.scalar(select(Region).where(Region.code == code))
        if existing is None:
            reg = Region(code=code, name=name)
            db.add(reg)
            db.flush()
            region_by_code[code] = reg
            result.regions_created += 1
        else:
            # Ad guncelle (yalnizca dolu ise ve degistiyse).
            if name and existing.name != name:
                existing.name = name
                result.regions_updated += 1
            region_by_code[code] = existing

    # 2) Line upsert (region+code bazli). Bransman 2. gecise birakilir.
    line_by_key: dict[str, Line] = {}   # "region|code" -> Line
    for key, pl in plan.lines.items():
        region = region_by_code.get(pl.region_code)
        if region is None:
            result.skipped += 1
            continue
        existing = db.scalar(
            select(Line).where(Line.region_id == region.id, Line.code == pl.code)
        )
        if existing is None:
            line = Line(region_id=region.id, code=pl.code, name=pl.name)
            db.add(line)
            db.flush()
            line_by_key[key] = line
            result.lines_created += 1
        else:
            if pl.name and existing.name != pl.name:
                existing.name = pl.name
                result.lines_updated += 1
            line_by_key[key] = existing

    # region|line_code -> Line (bransman cozumu icin de gerekli)
    line_by_region_code: dict[tuple[str, str], Line] = {}
    for key, line in line_by_key.items():
        rc, lc = key.split("|", 1)
        line_by_region_code[(rc, lc)] = line

    # 3) Pole upsert (line + sequence_no bazli). Pole id'lerini
    #    (region_code,line_code,seq) map'inde tut (ayni line code baska bolgede olabilir).
    pole_map: dict[tuple[str, str, int], Pole] = {}
    for pp in plan.poles:
        rc = pp.region_code
        line = line_by_region_code.get((rc, pp.line_code))
        if line is None:
            result.skipped += 1
            continue
        existing = db.scalar(
            select(Pole).where(Pole.line_id == line.id, Pole.sequence_no == pp.sequence_no)
        )
        if existing is None:
            pole = Pole(
                line_id=line.id,
                sequence_no=pp.sequence_no,
                name=pp.name,
                latitude=pp.latitude,
                longitude=pp.longitude,
                pole_type=pp.pole_type,
            )
            db.add(pole)
            db.flush()
            result.poles_created += 1
        else:
            existing.name = pp.name
            existing.latitude = pp.latitude
            existing.longitude = pp.longitude
            existing.pole_type = pp.pole_type
            pole = existing
            result.poles_updated += 1
        pole_map[(pp.region_code, pp.line_code, pp.sequence_no)] = pole

    # 4) Segment + cihaz. Cihaz, direk (seq) -> (seq+1) segmentine baglanir.
    #    Ayni slotta coklu cihaz: cihaz_sira'ya gore device_position_t yazilir
    #    (t = sira/(toplam+1) — esit dagilim, mevcut _slot_position formuluyle uyumlu).
    device_by_code = {d.code: d for d in db.scalars(select(Device)).all()}
    for pp in plan.poles:
        if not pp.devices:
            continue
        from_pole = pole_map.get((pp.region_code, pp.line_code, pp.sequence_no))
        to_pole = pole_map.get((pp.region_code, pp.line_code, pp.sequence_no + 1))
        if from_pole is None or to_pole is None:
            for pd in pp.devices:
                result.errors.append(
                    RowError(pd.excel_row, f"Cihaz için sonraki direk yok (segment kurulamadı): {pd.device_code}")
                )
            continue
        slot_total = len(pp.devices)
        # Sira'ya gore sirala (kucukten buyuge), t degerini ata.
        ordered = sorted(pp.devices, key=lambda d: d.cihaz_sira)
        for i, pd in enumerate(ordered, start=1):
            dev = device_by_code.get(pd.device_code)
            if dev is None:
                result.skipped += 1
                continue
            t = i / (slot_total + 1)  # esit dagilim; 1 cihaz -> 0.5
            already = db.scalar(select(LineSegment).where(LineSegment.device_id == dev.id))
            if already is not None:
                if already.from_pole_id == from_pole.id and already.to_pole_id == to_pole.id:
                    already.device_position_t = t
                    already.device_orientation = pd.orientation
                    continue
                result.errors.append(
                    RowError(pd.excel_row, f"Cihaz {pd.device_code} zaten bağlı — atlandı.")
                )
                continue
            seg = LineSegment(
                line_id=from_pole.line_id,
                from_pole_id=from_pole.id,
                to_pole_id=to_pole.id,
                device_id=dev.id,
                device_position_t=t,
                device_orientation=pd.orientation,
            )
            db.add(seg)
            db.flush()
            result.segments_created += 1

    # 5) Bransman (2. gecis) — tum hatlar olustuktan sonra.
    for key, pl in plan.lines.items():
        if not pl.branch_line_code or pl.branch_pole_seq is None:
            continue
        line = line_by_key.get(key)
        if line is None:
            continue
        # Bransmanin bagli olacagi hat: once plandaki direkler, yoksa DB.
        # DB fallback'i SIHIRBAZ icin sart: hedef hat plana hic girmez
        # (kullanici yalnizca yeni hatti gonderir), Excel'de de hedef hat
        # sayfada olmayabilir.
        target_pole = pole_map.get((pl.region_code, pl.branch_line_code, pl.branch_pole_seq))
        if target_pole is None:
            region = db.scalar(select(Region).where(Region.code == pl.region_code))
            if region is not None:
                t_line = db.scalar(
                    select(Line).where(
                        Line.region_id == region.id, Line.code == pl.branch_line_code
                    )
                )
                if t_line is not None:
                    target_pole = db.scalar(
                        select(Pole).where(
                            Pole.line_id == t_line.id,
                            Pole.sequence_no == pl.branch_pole_seq,
                        )
                    )
        if target_pole is not None:
            line.branched_from_pole_id = target_pole.id

    return result
