import io
import json
from datetime import datetime, timezone
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.enums import UserRole
from app.models.project_settings import ProjectSettings
from app.models.user import User
from app.schemas.event import SystemEventRead
from app.services.event_labels import (
    category_label,
    format_message,
    message_subject,
    severity_label,
    status_label,
)
from app.services.report_layout import format_report_time
from app.services.system_event_service import list_system_events

router = APIRouter(prefix="/events", tags=["events"])


#: OPERATOR'un denetim kaydinda GORMEMESI gereken kategoriler.
#:
#: `/events` ve `/events/export` yalnizca `get_current_user` ile korunuyordu,
#: yani rol ve kapsam kontrolu YOKTU: bir operator giris denemelerini, parola
#: sifirlamalarini, API anahtari uretimlerini ve kullanici yonetimi
#: hareketlerini okuyabiliyordu. Bunlar operatorun isi degil.
#:
#: Cihaz/alarm/ariza olaylari BILEREK acik kalir — operator kendi sahasinda
#: ne oldugunu gormeye devam eder, yani mevcut is akisi bozulmaz.
#:
#: NOT: olaylarin CIHAZ bazli kapsam suzgeci (operatorun yalnizca sorumlu
#: oldugu hatlarin cihaz olaylarini gormesi) BILINCLI olarak yapilmadi —
#: bugunku davranisi daraltmak bir urun karari; ayrica konusulmali.
_OPERATOR_GIZLI_KATEGORILER = frozenset({"auth", "security", "user"})


def _gizli_kategoriler(user: User) -> set[str] | None:
    """Bu kullanicinin gormemesi gereken olay kategorileri."""
    if user.role == UserRole.OPERATOR:
        return set(_OPERATOR_GIZLI_KATEGORILER)
    return None


# Export icin desteklenen formatlar. csv + json frontend tarafinda da
# uretilebilir ama endpoint backend-side fallback olarak duruyor (curl/script
# entegrasyonlari icin). xlsx + pdf sadece backend uzerinden.
_EXPORT_FORMATS = ("csv", "json", "xlsx", "pdf")


@router.get("", response_model=list[SystemEventRead])
def list_events(
    response: Response,
    category: str | None = Query(default=None),
    severity: str | None = Query(default=None),
    actor_username: str | None = Query(default=None),
    event_type: str | None = Query(default=None),
    event_type_like: str | None = Query(
        default=None,
        max_length=500,
        description="Virgulle ayrilmis ILIKE desenleri (OR) — orn. %_deleted,%_removed",
    ),
    device_code: str | None = Query(default=None, max_length=80),
    q: str | None = Query(default=None, max_length=200, description="Serbest metin arama"),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    limit: int = Query(default=1000, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Filtreli olay listesi (yeniden eskiye).

    Sayfalama: `limit` + `offset`; filtreye uyan TOPLAM kayit sayisi
    `X-Total-Count` header'inda doner (govde geriye uyumlu duz liste).
    Eski cagiranlar parametresiz kullanmaya devam edebilir (ilk 1000).
    """
    events, total = list_system_events(
        db,
        category=category,
        severity=severity,
        actor_username=actor_username,
        event_type=event_type,
        event_type_like=_split_patterns(event_type_like),
        device_code=device_code,
        q=q,
        date_from=date_from,
        date_to=date_to,
        exclude_categories=_gizli_kategoriler(user),
        limit=limit,
        offset=offset,
    )
    response.headers["X-Total-Count"] = str(total)
    return events


def _split_patterns(raw: str | None) -> list[str] | None:
    if not raw:
        return None
    patterns = [item.strip() for item in raw.split(",") if item.strip()]
    return patterns or None


def _format_rows_for_export(events, device_names: dict[str, str] | None = None) -> list[dict]:
    """Tum format'lar icin ortak satir yapisi — EKRANDAKI ile ayni icerik.

    Sutunlar ve degerler Olaylar tablosunun birebir karsiligi: cevrilmis
    mesaj/kategori/oncelik, olay tipinden turetilen DURUM ve cihaz KODU
    yerine cihaz ADI. Ham alanlar (event_type, metadata) ek kolon olarak
    kalir — makine tarafinda islenen CSV/JSON export'lari icin gerekli.
    """
    names = device_names or {}
    rows = []
    for ev in events:
        meta_summary = ""
        if ev.metadata_json:
            try:
                parsed = json.loads(ev.metadata_json)
                if isinstance(parsed, dict):
                    # `_i18n` teknik bir tas; ozet metinde kullaniciya gosterme.
                    meta_summary = ", ".join(
                        f"{k}={v}" for k, v in parsed.items() if k != "_i18n"
                    )
                else:
                    meta_summary = str(parsed)
            except (json.JSONDecodeError, TypeError):
                meta_summary = (ev.metadata_json or "")[:200]
        device_code = ev.device_code or ""
        # Cihaz kolonunun BOS kalmamasi icin geri kazanim: bazi olaylar
        # (ozellikle eski alarm kayitlari) cihazi yalnizca metadata'da tasir.
        meta_dict = None
        if not device_code and ev.metadata_json:
            try:
                meta_dict = json.loads(ev.metadata_json)
            except (json.JSONDecodeError, TypeError):
                meta_dict = None
            if isinstance(meta_dict, dict):
                device_code = str(
                    meta_dict.get("device_code") or meta_dict.get("device") or ""
                )
        ham = names.get(device_code)
        # Geriye uyum: harita degeri duz ad (str) da olabilir (testler/eski
        # cagiranlar) — (ad, seri) ciftine normallestir.
        ad_seri = (ham, "") if isinstance(ham, str) else ham
        rows.append({
            "id": ev.id,
            "created_at": format_report_time(ev.created_at, with_seconds=True),
            "severity": severity_label(ev.severity or ""),
            "category": category_label(ev.category or ""),
            "status": status_label(ev.event_type or ""),
            "message": message_subject(ev.message or "", ev.metadata_json),
            "message_full": format_message(ev.message or "", ev.metadata_json),
            "actor": ev.actor_username or "",
            "device": (ad_seri[0] if ad_seri else device_code),
            "serial": (ad_seri[1] if ad_seri else ""),
            "event_type": ev.event_type or "",
            "metadata": meta_summary,
        })
    return rows


def _device_name_map(db: Session) -> dict[str, tuple[str, str]]:
    """Cihaz kodu -> (ad, seri no). Arayuz tabloda ADI gosteriyor; export'ta
    ad + SERI NO birlikte iner (saha eslesmesi seri no ile yapiliyor)."""
    from app.models.device import Device

    rows = db.execute(select(Device.code, Device.name, Device.serial_number)).all()
    return {code: (name, serial or "") for code, name, serial in rows if code}


def _build_csv(rows: list[dict]) -> bytes:
    """RFC 4180 uyumlu CSV. UTF-8 BOM ile basliyor — Excel Turkce karakteri
    bozmadan acsin."""
    import csv

    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    csv_bytes = ("﻿" + buf.getvalue()).encode("utf-8")
    return csv_bytes


# Export kolon basliklari — ekrandaki sutun adlariyla ayni.
_COLUMN_TITLES = {
    "id": "Kayıt No",
    "created_at": "Tarih",
    "severity": "Öncelik",
    "category": "Kategori",
    "status": "Durum",
    "message": "Mesaj",
    "message_full": "Mesaj (tam)",
    "actor": "Kullanıcı",
    "device": "Cihaz",
    "serial": "Seri No",
    "event_type": "Olay Tipi",
    "metadata": "Ek Bilgi",
}


def _build_xlsx(rows: list[dict]) -> bytes:
    """openpyxl ile .xlsx — hucre tipleri korunur, baslik kalin, donduruldu."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Olaylar"

    if not rows:
        ws.append(["Filtreye uygun olay bulunamadı."])
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    headers = list(rows[0].keys())
    ws.append([_COLUMN_TITLES.get(h, h) for h in headers])
    # Header satiri: kalin + soluk turuncu arka plan (marka uyumu)
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FFFF8C00", end_color="FFFF8C00", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Sutun genisliklerini icerige gore otomatik ayarla
    col_widths = {h: max(len(_COLUMN_TITLES.get(h, h)), 12) for h in headers}
    for row in rows:
        for h in headers:
            v = str(row.get(h, ""))
            col_widths[h] = min(60, max(col_widths[h], len(v)))
    for i, h in enumerate(headers, start=1):
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(i)].width = col_widths[h] + 2

    # Ilk satiri dondur ki scroll yaparken baslik kalsin
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_pdf(rows: list[dict], *, settings_row: ProjectSettings | None) -> bytes:
    """A4 yatay olay raporu.

    Duzen: her sayfada solda EnerjiOne, sagda musteri logosu; altbilgide
    olusturma zamani + "Sayfa X / Y" (bkz. report_layout.ReportCanvas).

    Hucreler Paragraph — metin sutun genisliginde SARAR. Eski surum duz
    string kullaniyordu; uzun mesaj sutunu tasip komsu hucrenin uzerine
    biniyordu ("Toplu arsiv ayari..." metni Type sutununa giriyordu).
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="reportlab kutuphanesi yuklu degil; PDF export devre disi. "
            "Sistem yoneticisine 'pip install reportlab' calistirmasini soyleyin.",
        )

    from app.services.report_layout import (
        BRAND_ORANGE,
        FOOTER_HEIGHT,
        HEADER_HEIGHT,
        INK,
        ReportCanvas,
        decode_data_url_image,
        report_fonts,
    )

    regular_font, bold_font = report_fonts()
    customer_name = (settings_row.customer_name if settings_row else None) or ""
    project_name = (settings_row.project_name if settings_row else None) or ""
    customer_logo = decode_data_url_image(
        settings_row.customer_logo if settings_row else None
    )

    generated = format_report_time(datetime.now(timezone.utc), with_seconds=True)
    subtitle_parts = [part for part in (project_name, customer_name) if part]
    ReportCanvas.configure(
        title="Olay Raporu",
        subtitle=" · ".join(subtitle_parts),
        footer_left=f"Oluşturma: {generated}  ·  Toplam {len(rows)} olay",
        customer_logo=customer_logo,
        customer_name=customer_name,
    )

    margin = 12 * mm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        # Ustbilgi/altbilgi seritleri canvas'ta ciziliyor; govde onlarla
        # CAKISMASIN diye kenar bosluklari serit yuksekligi kadar buyuk.
        topMargin=margin + HEADER_HEIGHT,
        bottomMargin=margin + FOOTER_HEIGHT,
        title="EnerjiOne — Olay Raporu",
        author="EnerjiOne Grid",
    )

    cell_style = ParagraphStyle(
        "cell",
        fontName=regular_font,
        fontSize=7.5,
        leading=9.5,
        alignment=TA_LEFT,
        textColor=INK,
        wordWrap="CJK",  # uzun kesintisiz dizeleri (kod/yol) de kirar
    )
    head_style = ParagraphStyle(
        "cellHead",
        parent=cell_style,
        fontName=bold_font,
        fontSize=8,
        textColor=colors.white,
    )

    # Sutun seti EKRANDAKI ile ayni sirada: Tarih, Oncelik, Kategori, Mesaj,
    # Durum, Kullanici, Cihaz.
    headers = ["Tarih", "Öncelik", "Kategori", "Mesaj", "Durum", "Kullanıcı", "Cihaz", "Seri No"]
    col_widths = [28 * mm, 18 * mm, 28 * mm, 88 * mm, 24 * mm, 24 * mm, 34 * mm, 24 * mm]

    data: list[list] = [[Paragraph(h, head_style) for h in headers]]
    for row in rows:
        data.append([
            Paragraph(str(row.get("created_at", "")), cell_style),
            Paragraph(str(row.get("severity", "")), cell_style),
            Paragraph(str(row.get("category", "")), cell_style),
            Paragraph(str(row.get("message", "")), cell_style),
            Paragraph(str(row.get("status", "")), cell_style),
            Paragraph(str(row.get("actor", "")), cell_style),
            Paragraph(str(row.get("device", "")), cell_style),
            Paragraph(str(row.get("serial", "")), cell_style),
        ])

    story = []
    if len(data) > 1:
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_ORANGE),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
    else:
        story.append(
            Paragraph("Filtreye uygun olay bulunamadı.", cell_style)
        )

    doc.build(story, canvasmaker=ReportCanvas)
    return buf.getvalue()


# Export tavani: tek dosyada makul ust sinir (xlsx/pdf uretimi bellekte).
_EXPORT_MAX_ROWS = 20000


@router.get("/export")
def export_events(
    fmt: Literal["csv", "json", "xlsx", "pdf"] = Query("csv", description="Export format"),
    category: str | None = Query(default=None),
    severity: str | None = Query(default=None),
    actor_username: str | None = Query(default=None),
    event_type: str | None = Query(default=None),
    event_type_like: str | None = Query(default=None, max_length=500),
    device_code: str | None = Query(default=None, max_length=80),
    q: str | None = Query(default=None, max_length=200),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int | None = Query(default=None, ge=1, le=_EXPORT_MAX_ROWS),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Filtreli event listesini CSV/JSON/XLSX/PDF olarak indir.

    Frontend Events sayfasinda Export modal'i bu endpoint'i cagiriyor.
    Filtre query param'lari list_events ile birebir ayni — UI'da uygulanan
    filtreler ayni davranisi gosterir. Varsayilan sayfalama YOK: filtreye
    uyan tum kayitlar iner (ust sinir _EXPORT_MAX_ROWS). PDF'te yalnizca
    gorunen sayfa istenirse frontend offset/limit gonderir.
    """
    if fmt not in _EXPORT_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported format: {fmt!r}. Allowed: {', '.join(_EXPORT_FORMATS)}",
        )

    events, _total = list_system_events(
        db,
        category=category,
        severity=severity,
        actor_username=actor_username,
        event_type=event_type,
        event_type_like=_split_patterns(event_type_like),
        device_code=device_code,
        q=q,
        date_from=date_from,
        date_to=date_to,
        # Export listenin AYNI filtresini kullanmali; aksi halde ekranda
        # gizlenen olaylar CSV/PDF'ten sizardi.
        exclude_categories=_gizli_kategoriler(user),
        limit=min(limit or _EXPORT_MAX_ROWS, _EXPORT_MAX_ROWS),
        offset=offset,
    )
    rows = _format_rows_for_export(events, _device_name_map(db))
    now = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename_base = f"olaylar-{now}"

    if fmt == "csv":
        content = _build_csv(rows)
        return Response(
            content=content,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )
    if fmt == "json":
        content = json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")
        return Response(
            content=content,
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.json"'},
        )
    if fmt == "xlsx":
        content = _build_xlsx(rows)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )
    # pdf — musteri logosu/adi rapor basligi icin proje ayarlarindan gelir.
    settings_row = db.get(ProjectSettings, 1)
    content = _build_pdf(rows, settings_row=settings_row)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )
