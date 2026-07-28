import asyncio
import csv
import io
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import require_role, require_roles
from app.db.session import get_db
from app.models.device import Device
from app.models.enums import UserRole
from app.models.outbound_target import OutboundTarget, OutboundTopicMapping
from app.models.signal_catalog import SignalCatalog
from app.models.user import User
from app.schemas.modbus import ModbusPlanRead
from app.schemas.outbound import (
    OutboundTargetCreate,
    OutboundTargetRead,
    OutboundTargetUpdate,
    OutboundTopicMappingCreate,
    OutboundTopicMappingRead,
    OutboundTopicMappingUpdate,
)
from app.services import modbus_plan_service
from app.services.event_service import record_event
from app.services.iec104.bootstrap import redeploy_target
from app.services.iec104.registry import build_point_registry
from app.services.iec104.server import manager as iec104_manager

router = APIRouter(prefix="/outbound-targets", tags=["outbound-targets"])
logger = logging.getLogger(__name__)


def _refresh_mqtt_target(target_id: int) -> None:
    """MQTT target degisikligi sonrasi mqtt_publisher_service'i hot-reload et.

    Worker thread'leri DB'den fresh snapshot okumaz — explicit refresh
    cagirmak gerek. Refresh kendi icinde eski worker'i kapatip yenisini acar.
    """
    try:
        from app.services import mqtt_publisher_service
        mqtt_publisher_service.refresh_target(target_id)
    except Exception:  # noqa: BLE001
        logger.exception("mqtt_publisher_refresh_failed target_id=%s", target_id)


def _schedule_iec104_redeploy(db: Session, target_id: int) -> None:
    """Event loop'a threadsafe redeploy gorevini birakir.

    CRUD sync bir Session ile calisir ama IEC 104 server'i asyncio loop
    ustunde. `manager._loop` startup'ta bind edilmis olmali. Uygulama hic
    lifespan'a girmediyse sessizce geciyoruz (dev modunda HTTP client
    startup event'inden once istek atmadigi surece problem yok).
    """
    loop = iec104_manager._loop  # noqa: SLF001
    if loop is None:
        return

    async def _run() -> None:
        # Redeploy sirasinda yeni bir Session kullan, cunku gelen db thread-local.
        from app.db.session import SessionLocal

        scoped = SessionLocal()
        try:
            await redeploy_target(scoped, target_id)
        finally:
            scoped.close()

    try:
        asyncio.run_coroutine_threadsafe(_run(), loop)
    except RuntimeError:
        logger.debug("iec104_redeploy_schedule_failed target_id=%s", target_id)


@router.get("", response_model=list[OutboundTargetRead])
def list_outbound_targets(
    _: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    from sqlalchemy.orm import selectinload
    stmt = (
        select(OutboundTarget)
        .options(selectinload(OutboundTarget.topic_mappings))
        .order_by(OutboundTarget.name.asc())
    )
    return list(db.scalars(stmt).unique().all())


@router.get("/runtime-status")
def outbound_runtime_status(
    _: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
):
    """UI 'Durum' sutunu icin canli rozet verisi.

    Donen yapi: {target_id: {protocol_specific_fields}}.
    - MQTT: connected, last_publish_at, sent_total, failed_total
    - REST/MQTT_dispatch: last_success_at, last_failure_at, last_error, sent_total, failed_total
    """
    from app.services import mqtt_publisher_service
    from app.services.outbound_dispatch_service import delivery_status_snapshot

    mqtt_runtime = mqtt_publisher_service.runtime_status()
    delivery = delivery_status_snapshot()
    out: dict[int, dict] = {}
    for tid, st in mqtt_runtime.items():
        out.setdefault(tid, {}).update(st)
    for tid, st in delivery.items():
        out.setdefault(tid, {}).update(st)
    return out


@router.get("/{target_id}/auto-topics")
def outbound_auto_topics(
    target_id: int,
    _: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    """MQTT target icin uretilecek otomatik topic listesi.

    Devices tablosundan tum cihazlari okur, target'in template'ine
    gore her cihaz icin olusacak topic'leri donderir.
    """
    target = db.scalar(select(OutboundTarget).where(OutboundTarget.id == target_id))
    if target is None:
        raise HTTPException(status_code=404, detail="Target bulunamadi")
    if target.protocol != "mqtt":
        raise HTTPException(status_code=400, detail="Bu endpoint sadece MQTT target icin")

    devices = list(db.scalars(select(Device).order_by(Device.code.asc())).all())
    device_dicts = [{"code": d.code, "name": d.name} for d in devices]

    from app.services import mqtt_publisher_service
    rows = mqtt_publisher_service.auto_topics_for_target(target_id, device_dicts)
    return {"target_id": target_id, "device_count": len(device_dicts), "topics": rows}


@router.post("", response_model=OutboundTargetRead, status_code=status.HTTP_201_CREATED)
def create_outbound_target(
    payload: OutboundTargetCreate,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    # Hedef adi PROTOKOLDEN BAGIMSIZ benzersizdir. Eskiden mesaj sadece
    # "already exists" diyordu; operator Modbus hedefi eklerken ayni ada
    # sahip bir IEC 104 hedefi oldugunu anlamiyor, sistemin protokolleri
    # karistirdigini saniyordu. Mesaj artik carpisan kaydin protokolunu de
    # soyluyor.
    existing = db.scalar(select(OutboundTarget).where(OutboundTarget.name == payload.name))
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"'{payload.name}' adinda bir outbound hedef zaten var "
                f"({existing.protocol.upper()}). Hedef adlari protokolden "
                f"bagimsiz olarak benzersiz olmali — farkli bir ad girin."
            ),
        )
    row = OutboundTarget(**payload.model_dump())
    db.add(row)
    db.flush()
    record_event(
        db,
        category="outbound",
        event_type="outbound_created",
        severity="info",
        actor_username=current_user.username,
        message=f"Outbound hedef eklendi: {row.name} ({row.protocol.upper()})",
        metadata={"target_id": row.id, "protocol": row.protocol},
        i18n_key="outbound_created",
        i18n_params={"name": row.name, "protocol": row.protocol.upper()},
    )
    db.commit()
    db.refresh(row)
    if row.protocol == "iec104":
        _schedule_iec104_redeploy(db, row.id)
    elif row.protocol == "mqtt":
        _refresh_mqtt_target(row.id)
    return row


@router.post("/{target_id}/test")
def test_outbound_target(
    target_id: int,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
) -> dict:
    """REST/webhook hedefine temsili bir olay POST'lar (n8n vb. dogrulamasi).

    Kaydedilmis hedefin endpoint + auth_header/token ayarlarini kullanarak
    ornek bir alarm event JSON'i gonderir; n8n'de "Listen for test event" ile
    akis kurmak icin idealdir. Sonuc {ok, detail} doner ve 'Durum' sutununu
    gunceller. Sadece REST protokolu desteklenir (MQTT/IEC104 haric)."""
    from uuid import uuid4

    from app.services.outbound_dispatch_service import _record_delivery, _send_rest

    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    if target.protocol != "rest":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Test yalnizca REST/webhook hedefleri icin desteklenir.",
        )

    sample = {
        "message_id": str(uuid4()),
        "event_kind": "alarm",
        "test": True,
        "device_code": "DEV-001",
        "device_name": "Ornek Fider",
        "signal_key": "sat01.current_phase_a",
        "value": 245.7,
        "quality": "good",
        "severity": "critical",
        "message": "EnerjiOne Grid test webhook event",
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    try:
        _send_rest(target, sample)
        _record_delivery(target.id, ok=True)
        return {"ok": True, "detail": f"Test verisi gonderildi -> {target.endpoint}"}
    except Exception as exc:  # noqa: BLE001
        _record_delivery(target.id, ok=False, error=str(exc))
        return {"ok": False, "detail": str(exc)}


@router.patch("/{target_id}", response_model=OutboundTargetRead)
def update_outbound_target(
    target_id: int,
    payload: OutboundTargetUpdate,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    row = db.get(OutboundTarget, target_id)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    updates = payload.model_dump(exclude_unset=True)
    for key, value in updates.items():
        setattr(row, key, value)
    record_event(
        db,
        category="outbound",
        event_type="outbound_updated",
        severity="info",
        actor_username=current_user.username,
        message=f"Outbound target updated: {row.name}",
        metadata={"target_id": row.id, "fields": list(updates.keys())},
        i18n_key="outbound_updated",
        i18n_params={"name": row.name},
    )
    db.commit()
    db.refresh(row)
    if row.protocol == "iec104":
        _schedule_iec104_redeploy(db, row.id)
    elif row.protocol == "mqtt":
        _refresh_mqtt_target(row.id)
    return row


@router.delete("/{target_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_outbound_target(
    target_id: int,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    row = db.get(OutboundTarget, target_id)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    was_iec104 = row.protocol == "iec104"
    saved_id = row.id
    name = row.name
    proto = row.protocol
    db.delete(row)
    record_event(
        db,
        category="outbound",
        event_type="outbound_deleted",
        severity="warning",
        actor_username=current_user.username,
        message=f"Outbound hedef silindi: {name} ({proto.upper()})",
        metadata={"target_id": saved_id, "protocol": proto},
        i18n_key="outbound_deleted",
        i18n_params={"name": name, "protocol": proto.upper()},
    )
    db.commit()
    if was_iec104:
        # undeploy ayni IEC 104 manager uzerinde calisir; schedule threadsafe.
        loop = iec104_manager._loop  # noqa: SLF001
        if loop is not None:
            asyncio.run_coroutine_threadsafe(iec104_manager.undeploy(saved_id), loop)
    elif proto == "mqtt":
        # MQTT worker'i temizle — refresh_target target yoksa sadece eski
        # worker'i durdurur (silinmis target icin yeni acmaz).
        _refresh_mqtt_target(saved_id)
    return None


# ----- IEC 104 point list CSV -----

# IEC 104 ASDU Type ID -> kisa kod ve ad. SCADA mühendisleri CSV'de bu adi tanir.
_IEC104_TYPE_LABEL: dict[int, str] = {
    1: "M_SP_NA_1",
    3: "M_DP_NA_1",
    9: "M_ME_NA_1",
    11: "M_ME_NB_1",
    13: "M_ME_NC_1",
    15: "M_IT_NA_1",
    30: "M_SP_TB_1",
    31: "M_DP_TB_1",
    34: "M_ME_TD_1",
    35: "M_ME_TE_1",
    36: "M_ME_TF_1",
    37: "M_IT_TB_1",
}


@router.get("/{target_id}/iec104-runtime")
def get_iec104_runtime(
    target_id: int,
    _: User = Depends(
        require_roles([UserRole.ENGINEER, UserRole.INSTALLER])
    ),
    db: Session = Depends(get_db),
) -> dict:
    """Bu IEC 104 hedefinin canli durumunu doner.

    Gercek IEC 104 TCP server'lari ayri `iec104-outbound` Docker servisinde
    calisir. Bu endpoint hedef DB satirini + iec104-outbound HTTP health
    endpoint'inden runtime'i toplar. Eski (gomulu) backend manager'i fallback
    olarak kalir — ama compose'da artik kullanilmiyor."""
    import os
    import requests as _requests

    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    if target.protocol != "iec104":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Yalniz IEC 104 hedefler icin runtime sorgulanir.",
        )
    raw = (target.iec104_allowed_peers or "").strip()
    allowed = [p.strip() for p in raw.split(",") if p.strip()]

    # Once compose servisinden cek (asil server orada)
    server_running = False
    connected_clients: list[dict] = []
    base = os.getenv("IEC104_OUTBOUND_HEALTH_URL", "http://iec104-outbound:8014")
    try:
        resp = _requests.get(f"{base}/runtime/{target_id}", timeout=2.5)
        if resp.status_code == 200:
            data = resp.json() or {}
            server_running = bool(data.get("server_running"))
            connected_clients = list(data.get("connected_clients") or [])
        else:
            # Servis cevap verdi ama hata; gomulu manager'a dus.
            server_running = iec104_manager.is_running(target.id)
            connected_clients = iec104_manager.connected_clients(target.id)
    except Exception:
        # Servis erisilemiyor — gomulu manager (eski mod) ile devam et.
        server_running = iec104_manager.is_running(target.id)
        connected_clients = iec104_manager.connected_clients(target.id)

    return {
        "target_id": target.id,
        "server_running": server_running,
        "whitelist_active": bool(allowed),
        "allowed_peers": allowed,
        "connected_clients": connected_clients,
    }


@router.get("/{target_id}/iec104-points.csv")
def export_iec104_points_csv(
    target_id: int,
    _: User = Depends(
        require_roles([UserRole.ENGINEER, UserRole.INSTALLER])
    ),
    db: Session = Depends(get_db),
) -> Response:
    """Bu IEC 104 hedefine ait tum cihaz x sinyal kombinasyonlarinin point list'ini
    CSV olarak doner.

    Saha mühendisi SCADA tarafindaki IEC 104 master'a noktalari (CA, IOA, Type ID,
    sinyal etiketi) bu CSV'den toplu olarak girebilsin diye duz format:

      device_code, device_name, ca, ioa, type_id, type_code, signal_key, label, unit, scale, offset

    Yalniz `protocol=iec104` hedefler icin gecerli; `is_active=False` cihazlar ve
    `iec104_enabled=False` veya `iec104_type_id` bos sinyaller atlanir (yayinda
    olanla CSV ozdes).
    """
    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    if target.protocol != "iec104":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu hedef IEC 104 degil; point list yalnizca IEC 104 hedefler icin uretilir.",
        )

    devices = list(db.scalars(select(Device).order_by(Device.code)).all())
    signals = list(
        db.scalars(
            select(SignalCatalog)
            .where(SignalCatalog.is_active.is_(True))
            .order_by(SignalCatalog.dnp3_object_group, SignalCatalog.dnp3_index)
        ).all()
    )

    default_ca = target.iec104_common_address or 1
    registry = build_point_registry(
        target_id=target.id,
        default_common_address=default_ca,
        devices=devices,
        signals=signals,
    )

    # SignalCatalog lookup signal_key -> meta (label, unit, scale, offset).
    sig_meta: dict[str, SignalCatalog] = {s.key: s for s in signals}
    dev_meta: dict[str, Device] = {d.code: d for d in devices}

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "device_code",
            "device_name",
            "common_address",
            "ioa",
            "type_id",
            "type_code",
            "signal_key",
            "label",
            "unit",
            "scale",
            "offset",
        ]
    )
    # CA, IOA artan sira — SCADA tarafinda goz tarafindan kontrol kolay.
    for p in sorted(registry.points, key=lambda x: (x.common_address, x.ioa)):
        sig = sig_meta.get(p.signal_key)
        dev = dev_meta.get(p.device_code)
        writer.writerow(
            [
                p.device_code,
                dev.name if dev else "",
                p.common_address,
                p.ioa,
                p.type_id,
                _IEC104_TYPE_LABEL.get(p.type_id, ""),
                p.signal_key,
                sig.label if sig else "",
                sig.unit if sig and sig.unit else "",
                sig.scale if sig else "",
                sig.offset if sig else "",
            ]
        )

    csv_text = buf.getvalue()
    # Excel UTF-8 BOM ile dogru gostersin
    body = ("﻿" + csv_text).encode("utf-8")
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in target.name)
    filename = f"iec104-points-{safe_name}-{ts}.csv"
    return Response(
        content=body,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Point-Count": str(len(registry.points)),
            "Access-Control-Expose-Headers": "Content-Disposition, X-Point-Count",
        },
    )


@router.get("/{target_id}/iec104-points.xlsx")
def export_iec104_points_xlsx(
    target_id: int,
    _: User = Depends(
        require_roles([UserRole.ENGINEER, UserRole.INSTALLER])
    ),
    db: Session = Depends(get_db),
) -> Response:
    """SCADA mühendisi için Excel (.xlsx) point list.

    İçerik:
      - Sheet 'Points' — her cihaz × aktif IEC 104 sinyali (CA, IOA, Type ID,
        type_code, etiket, birim, scale, offset)
      - Sheet 'Devices' — cihaz listesi: cihaz adı/kodu + ASDU CA (default
        miydi gösterir)

    Sadece `iec104_enabled=True` ve `iec104_type_id` dolu sinyaller dahil.
    Cihaz `iec104_common_address` boş ise default kullanılır ve 'is_default'
    kolonu işaretlenir."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl kurulu değil. requirements.txt'i güncelleyip yeniden başlatın.",
        )

    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    if target.protocol != "iec104":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Yalnız IEC 104 hedefler için point list üretilir.",
        )

    devices = list(db.scalars(select(Device).order_by(Device.code)).all())
    signals = list(
        db.scalars(
            select(SignalCatalog)
            .where(SignalCatalog.is_active.is_(True))
            .order_by(SignalCatalog.dnp3_object_group, SignalCatalog.dnp3_index)
        ).all()
    )
    default_ca = target.iec104_common_address or 1
    registry = build_point_registry(
        target_id=target.id,
        default_common_address=default_ca,
        devices=devices,
        signals=signals,
    )
    sig_meta: dict[str, SignalCatalog] = {s.key: s for s in signals}
    dev_meta: dict[str, Device] = {d.code: d for d in devices}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Points
    ws = wb.active
    ws.title = "Points"
    headers = [
        "Sıra", "Cihaz Adı", "Cihaz Kodu", "ASDU CA", "IOA",
        "Type ID", "Type Code", "Sinyal Anahtarı", "Etiket", "Birim", "Scale", "Offset",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sorted_points = sorted(registry.points, key=lambda p: (p.common_address, p.ioa))
    for row_idx, p in enumerate(sorted_points, start=1):
        sig = sig_meta.get(p.signal_key)
        dev = dev_meta.get(p.device_code)
        ws.append([
            row_idx,
            dev.name if dev else "",
            p.device_code,
            p.common_address,
            p.ioa,
            p.type_id,
            _IEC104_TYPE_LABEL.get(p.type_id, ""),
            p.signal_key,
            sig.label if sig else "",
            sig.unit if sig and sig.unit else "",
            sig.scale if sig else "",
            sig.offset if sig else "",
        ])
    # Kolon genislikleri (yaklasik)
    widths = [6, 24, 16, 10, 10, 10, 14, 32, 30, 8, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # Sheet 2: Devices
    ws2 = wb.create_sheet("Devices")
    dev_headers = ["Sıra", "Cihaz Adı", "Cihaz Kodu", "ASDU CA", "Default mi?"]
    ws2.append(dev_headers)
    for col_idx in range(1, len(dev_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row_idx, d in enumerate(devices, start=1):
        own_ca = getattr(d, "iec104_common_address", None)
        is_default = own_ca is None
        effective_ca = own_ca if own_ca is not None else default_ca
        ws2.append([row_idx, d.name, d.code, effective_ca, "Evet" if is_default else "Hayır"])
    for i, w in enumerate([6, 28, 18, 10, 12], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in target.name)
    filename = f"iec104-points-{safe_name}-{ts}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Point-Count": str(len(sorted_points)),
            "Access-Control-Expose-Headers": "Content-Disposition, X-Point-Count",
        },
    )


def _require_modbus_target(db: Session, target_id: int) -> OutboundTarget:
    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found"
        )
    if target.protocol != "modbus":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu hedef Modbus degil; adres plani yalnizca Modbus hedefler icin uretilir.",
        )
    return target


@router.get("/{target_id}/modbus-plan", response_model=ModbusPlanRead)
def get_modbus_plan(
    target_id: int,
    _: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    """Bu Modbus hedefinin adres plani: cihaz slotlari + tam adres tablosu.

    Plan `modbus_plan_service` tarafindan uretilir; worker da AYNI fonksiyonun
    ciktisini kullanir, yani burada gorunen adres sahada yayinlanan adrestir.
    Cagri sirasinda eksik cihaz slotlari otomatik atanir ve kalici yazilir
    (mevcut slotlar korunur — adresler kaymaz).
    """
    target = _require_modbus_target(db, target_id)
    return ModbusPlanRead(**modbus_plan_service.serialize_plan(db, target))


@router.get("/{target_id}/modbus-points.csv")
def export_modbus_points_csv(
    target_id: int,
    _: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
) -> Response:
    """Adres tablosunu CSV olarak indir — SCADA'ya toplu tag girisi icin.

    Sutunlar: device_code, device_name, unit_id, function, function_name,
    address, modicon (40001 tarzi klasik gosterim), signal_key, label, source,
    data_type, unit, word_count, scale, offset
    """
    target = _require_modbus_target(db, target_id)
    plan = modbus_plan_service.serialize_plan(db, target)

    fn_names = {1: "coil", 2: "discrete_input", 3: "holding_register", 4: "input_register"}
    # Klasik "Modicon" gosterimi: coil 1-tabanli 00001, discrete 10001,
    # input register 30001, holding register 40001.
    fn_bases = {1: 1, 2: 10_001, 3: 40_001, 4: 30_001}

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "device_code", "device_name", "unit_id", "function", "function_name",
            "address", "modicon", "signal_key", "label", "source", "data_type",
            "unit", "word_count", "scale", "offset",
        ]
    )
    for p in plan["points"]:
        writer.writerow(
            [
                p["device_code"], p["device_name"], p["unit_id"], p["function"],
                fn_names.get(p["function"], str(p["function"])),
                p["address"], fn_bases.get(p["function"], 0) + p["address"],
                p["signal_key"], p["label"], p["source"], p["data_type"],
                p["unit"] or "", p["word_count"], p["scale"], p["offset"],
            ]
        )

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    safe_name = "".join(c if c.isalnum() or c in "-_" else "-" for c in target.name)
    filename = f"modbus-points-{safe_name}-{stamp}.csv"
    # UTF-8 BOM: Excel'in Turkce karakterleri dogru acmasi icin.
    return Response(
        content="﻿" + buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Point-Count": str(len(plan["points"])),
            "Access-Control-Expose-Headers": "Content-Disposition, X-Point-Count",
        },
    )


@router.post("/{target_id}/auto-assign-device-ca", response_model=dict)
def auto_assign_device_ca(
    target_id: int,
    payload: dict | None = None,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    """Bu hedef altinda CA'si bos cihazlara sirayla 1, 2, 3... atar.

    payload (opsiyonel):
      start_at: int — baslangic (default 1)
      overwrite: bool — mevcut CA'leri de ezerek yenile (default False)

    Cikti: {assigned: N, skipped: M, devices: [{code, ca}]}
    """
    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    if target.protocol != "iec104":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Auto-assign yalnız IEC 104 hedefler için.",
        )
    payload = payload or {}
    start_at = int(payload.get("start_at", 1))
    overwrite = bool(payload.get("overwrite", False))
    if start_at < 0 or start_at > 65000:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="start_at 0-65000 arası olmalı.")

    devices = list(db.scalars(select(Device).order_by(Device.code)).all())
    used_cas: set[int] = set()
    if not overwrite:
        for d in devices:
            ca = getattr(d, "iec104_common_address", None)
            if ca is not None:
                used_cas.add(int(ca))

    assigned = 0
    skipped = 0
    next_ca = start_at
    result_list: list[dict] = []
    for d in devices:
        existing = getattr(d, "iec104_common_address", None)
        if not overwrite and existing is not None:
            skipped += 1
            continue
        # Bir sonraki bos CA bul
        while next_ca in used_cas:
            next_ca += 1
            if next_ca > 65534:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Atanabilir CA aralığı tükendi.",
                )
        d.iec104_common_address = next_ca
        used_cas.add(next_ca)
        result_list.append({"code": d.code, "ca": next_ca})
        assigned += 1
        next_ca += 1

    if assigned > 0:
        record_event(
            db,
            category="outbound_target",
            event_type="iec104_auto_assign_ca",
            severity="info",
            actor_username=current_user.username,
            message=(
                f"'{target.name}' hedefi için {assigned} cihaza IEC 104 Common Address "
                f"otomatik atandı (başlangıç: {start_at}, atlanan: {skipped}, overwrite: {overwrite})"
            ),
            metadata={
                "target_id": target.id,
                "target_name": target.name,
                "assigned": assigned,
                "skipped": skipped,
                "start_at": start_at,
                "overwrite": overwrite,
            },
        )
    db.commit()
    # IEC 104 server'i yeniden deploy et — CA degisikligi nokta haritasini etkiler.
    _schedule_iec104_redeploy(db, target.id)
    return {"assigned": assigned, "skipped": skipped, "devices": result_list}


# ===========================================================================
# MQTT Custom Topic Mapping CRUD
# ===========================================================================
#
# Her MQTT outbound target icin operator UI'da "Custom Topic Mapping" modal
# acabilir; cihaz + sinyal listesi -> ozel topic eslestirmesi yapar. Default
# template (`{prefix}/{customer}/{device}/{source}/{datatype}/telemetry`)
# bypass edilir. Birden fazla mapping ayni reading'i yakalarsa hepsine
# publish edilir.
#
# Routes:
#   GET    /outbound-targets/{tid}/topic-mappings        -> list
#   POST   /outbound-targets/{tid}/topic-mappings        -> create
#   PATCH  /outbound-targets/{tid}/topic-mappings/{mid}  -> update
#   DELETE /outbound-targets/{tid}/topic-mappings/{mid}  -> delete

def _ensure_mqtt_target(db: Session, target_id: int) -> OutboundTarget:
    target = db.get(OutboundTarget, target_id)
    if target is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outbound target not found")
    if target.protocol != "mqtt":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Topic mapping yalniz MQTT target'lar icin gecerli.",
        )
    return target


@router.get(
    "/{target_id}/topic-mappings",
    response_model=list[OutboundTopicMappingRead],
)
def list_topic_mappings(
    target_id: int,
    _: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    _ensure_mqtt_target(db, target_id)
    return list(
        db.scalars(
            select(OutboundTopicMapping)
            .where(OutboundTopicMapping.target_id == target_id)
            .order_by(OutboundTopicMapping.id.asc())
        ).all()
    )


@router.post(
    "/{target_id}/topic-mappings",
    response_model=OutboundTopicMappingRead,
    status_code=status.HTTP_201_CREATED,
)
def create_topic_mapping(
    target_id: int,
    payload: OutboundTopicMappingCreate,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    target = _ensure_mqtt_target(db, target_id)
    row = OutboundTopicMapping(target_id=target_id, **payload.model_dump())
    db.add(row)
    db.flush()
    record_event(
        db,
        category="outbound",
        event_type="mqtt_topic_mapping_created",
        severity="info",
        actor_username=current_user.username,
        message=f"MQTT topic mapping eklendi: {target.name} -> {row.topic}",
        metadata={
            "target_id": target_id,
            "mapping_id": row.id,
            "topic": row.topic,
            "device_codes": row.device_codes,
            "signal_keys": row.signal_keys,
        },
    )
    db.commit()
    db.refresh(row)
    _refresh_mqtt_target(target_id)
    return row


@router.patch(
    "/{target_id}/topic-mappings/{mapping_id}",
    response_model=OutboundTopicMappingRead,
)
def update_topic_mapping(
    target_id: int,
    mapping_id: int,
    payload: OutboundTopicMappingUpdate,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    _ensure_mqtt_target(db, target_id)
    row = db.get(OutboundTopicMapping, mapping_id)
    if row is None or row.target_id != target_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Topic mapping not found")
    updates = payload.model_dump(exclude_unset=True)
    for key, value in updates.items():
        setattr(row, key, value)
    record_event(
        db,
        category="outbound",
        event_type="mqtt_topic_mapping_updated",
        severity="info",
        actor_username=current_user.username,
        message=f"MQTT topic mapping guncellendi: {row.topic}",
        metadata={
            "target_id": target_id,
            "mapping_id": mapping_id,
            "fields": list(updates.keys()),
        },
    )
    db.commit()
    db.refresh(row)
    _refresh_mqtt_target(target_id)
    return row


@router.delete(
    "/{target_id}/topic-mappings/{mapping_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def delete_topic_mapping(
    target_id: int,
    mapping_id: int,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    _ensure_mqtt_target(db, target_id)
    row = db.get(OutboundTopicMapping, mapping_id)
    if row is None or row.target_id != target_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Topic mapping not found")
    topic_snapshot = row.topic
    db.delete(row)
    record_event(
        db,
        category="outbound",
        event_type="mqtt_topic_mapping_deleted",
        severity="warning",
        actor_username=current_user.username,
        message=f"MQTT topic mapping silindi: {topic_snapshot}",
        metadata={"target_id": target_id, "mapping_id": mapping_id, "topic": topic_snapshot},
    )
    db.commit()
    _refresh_mqtt_target(target_id)
    return None


# ===========================================================================
# MQTT TLS Sertifika upload
# ===========================================================================
#
# Operator UI'da TLS toggle'i acinca 3 sertifika dosyasi (CA, client cert,
# client key) upload edebilir. Backend bunlari host'ta sabit dizine yazar
# ve target'in mqtt_tls_*_path alanini otomatik gunceller. Operator path
# manuel yazmak zorunda kalmaz.
#
# Dosya saklama dizini: get_backup_dir().parent / "mqtt-certs" / "{target_id}"
# Yani container'da /var/lib/e1-backups/../mqtt-certs/{id}/  = /var/lib/mqtt-certs
# Volume mount disinda — sadece backend container'da gecerli. Container
# yeniden olusturulursa kaybolur; operator yeniden upload eder. Production
# icin docker-compose'da bind mount eklenebilir.

# Cert tipleri ile model field mapping
_CERT_KINDS = {
    "ca": "mqtt_tls_ca_path",
    "cert": "mqtt_tls_cert_path",
    "key": "mqtt_tls_key_path",
}

# Sertifika dizini — backup-data volume'unun parent'i icinde sabit dizin.
# /var/lib/e1-backups -> parent /var/lib -> mqtt-certs.
def _mqtt_cert_dir(target_id: int) -> Path:
    from app.services.backup_service import get_backup_dir
    root = Path(os.getenv("MQTT_CERT_DIR", str(get_backup_dir().parent / "mqtt-certs")))
    target_dir = root / str(target_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


@router.post("/{target_id}/mqtt-cert/{kind}", response_model=OutboundTargetRead)
async def upload_mqtt_cert(
    target_id: int,
    kind: str,
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    """MQTT TLS sertifikasi upload — CA / cert / key.

    Yuklenen dosya disk'e yazilir, target'in ilgili mqtt_tls_*_path alani
    otomatik atanir. mqtt_publisher_service hot-reload tetiklenir.

    Boyut limiti: 100 KB (PEM/CRT/KEY dosyalari tipik 1-10 KB).
    """
    target = _ensure_mqtt_target(db, target_id)
    if kind not in _CERT_KINDS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gecersiz cert tipi: {kind}. Gecerli: ca, cert, key",
        )
    field_name = _CERT_KINDS[kind]

    # Dosyayi oku + boyut check
    content = await file.read()
    if len(content) > 100 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Sertifika dosyasi cok buyuk ({len(content)} byte). Max 100 KB.",
        )
    # PEM kontrolu — basit (BEGIN/END marker arariz)
    text_preview = content[:200].decode("ascii", errors="ignore")
    if "-----BEGIN" not in text_preview:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Dosya PEM formatinda gorunmuyor (BEGIN marker yok). .crt/.pem/.key bekleniyor.",
        )

    # Dosyayi yaz
    cert_dir = _mqtt_cert_dir(target_id)
    ext = {"ca": "ca.crt", "cert": "client.crt", "key": "client.key"}[kind]
    target_path = cert_dir / ext
    try:
        target_path.write_bytes(content)
        os.chmod(target_path, 0o600)  # key dosyasi sadece backend user okuyabilsin
    except OSError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Sertifika kaydedilemedi: {exc}",
        )

    # Target field'ini guncelle
    setattr(target, field_name, str(target_path))
    record_event(
        db,
        category="outbound",
        event_type="mqtt_cert_uploaded",
        severity="info",
        actor_username=current_user.username,
        message=f"MQTT TLS sertifikasi yuklendi: {target.name} ({kind})",
        metadata={
            "target_id": target_id,
            "kind": kind,
            "path": str(target_path),
            "size": len(content),
            "filename": file.filename,
        },
    )
    db.commit()
    db.refresh(target)
    _refresh_mqtt_target(target_id)
    return target


@router.delete("/{target_id}/mqtt-cert/{kind}", response_model=OutboundTargetRead)
def delete_mqtt_cert(
    target_id: int,
    kind: str,
    current_user: User = Depends(require_roles([UserRole.INSTALLER, UserRole.ENGINEER])),
    db: Session = Depends(get_db),
):
    """Yuklenen MQTT TLS sertifikasini sil — disk + target field NULL."""
    target = _ensure_mqtt_target(db, target_id)
    if kind not in _CERT_KINDS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gecersiz cert tipi: {kind}",
        )
    field_name = _CERT_KINDS[kind]
    current_path = getattr(target, field_name)
    if current_path:
        try:
            Path(current_path).unlink(missing_ok=True)
        except Exception:  # noqa: BLE001
            logger.exception("mqtt_cert_unlink_failed path=%s", current_path)
    setattr(target, field_name, None)
    record_event(
        db,
        category="outbound",
        event_type="mqtt_cert_deleted",
        severity="warning",
        actor_username=current_user.username,
        message=f"MQTT TLS sertifikasi silindi: {target.name} ({kind})",
        metadata={"target_id": target_id, "kind": kind},
    )
    db.commit()
    db.refresh(target)
    _refresh_mqtt_target(target_id)
    return target
